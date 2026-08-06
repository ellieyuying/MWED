#!/usr/bin/env python3
"""
短语级 CEFR 词汇复杂度指标计算（sense-aware）。

吃 run_pipeline 产出的 hit_detail.xlsx，按作文(answer)聚合出 16 个短语级指标，
并计算与写作总分(1-40)的 Pearson 相关。

等级取值（sense-aware）：
  单义短语用其固定 CEFR；多义短语用【消歧义项】在多义表里的真实等级
  （按 词条+消歧义项 查 等级）；not_applied / 解析失败 / 无法映射的义项
  无有效等级，不计入等级类指标（但仍计入 pv_ratio / multi_ratio 分母）。

指标：
  比例类: PH_ratio_{A1,A2,B1,B2,C1,C2,A,B,C,AboveB1,AboveB2}
  均值类: PH_mean_score, PH_high_low_ratio, PH_score_var
  短语特有: PH_pv_ratio(短语动词比例), PH_multi_ratio(多义短语比例)

用法：
  # 逐篇作文（默认，N≈2470）
  python compute_phrase_indices.py --hit-detail 结果_v12/hit_detail.xlsx \
      --ambig phrase_extraction/data/多义短语_v10.xlsx \
      --out-indices 结果_v12/phrase_indices_v12.csv --out-corr 结果_v12/phrase_corr_v12.csv
  # 按试卷聚合（N=1236，D4 稳健性；默认输出落到 *_perscript.csv）
  python compute_phrase_indices.py --unit script

================================================================================
与 Hu et al. (2025, Behav Res Methods 57:226) 的 4 处口径差异——刻意为之，务必在
简报/论文的“边界”里显式声明，否则“对标 Hu et al. 词级指标”的说法不成立。
--------------------------------------------------------------------------------
D1  未用移动平均(MA)。论文 Type I/II 全是 MA_X：滑动窗口 n=100、逐 token 移动、
    每窗算比例再取均值，目的是抵消文本长度影响。本脚本用【全局简单比例】count/tot。
    理由：每篇仅约 9 个短语命中，无法在其上滑 100-token 窗；短语稀疏使 MA 不适用。
    代价：这些 PH_ratio_* 不是论文的 MA_X，且未做 type/token 之分（论文发现 type 版更强）。
D2  两套分母。等级分布类(PH_ratio_*/mean_score/score_var/high_low_ratio)分母=有等级命中数；
    短语属性类(PH_pv_ratio/PH_multi_ratio)分母=全部命中。后者【必须】含未定级命中——
    未消歧成功的多义命中本身就是多义，否则 multi 系统性漏计。已输出 PH_level_coverage
    (=有等级命中/全部命中) 记录被前者排除的比例。见 _essay_indices() 注释。
D3  模式不对齐。论文旗舰词级结果用 lazyA1 模式；本项目短语用 sense-aware(≈论文 original)，
    而词级基线(见 核查说明/简报)用 Min/Lowest 模式。故 ΔR² 是两种等级赋值哲学相减，
    增量可能偏乐观——严格版需两侧同模式重跑。
D4  分析单位。论文 N=1236 试卷，一卷一个总分(含两 task)。本脚本默认按 answer 聚合
    (N≈2470)，同卷两篇【共享】总分→违背独立性、虚高 p 值。用 --unit script 可复现
    论文单位(N=1236)作稳健性旁证：相关不降反升(稀疏噪声被两篇合并稀释)。
================================================================================
"""
import argparse
from collections import defaultdict, Counter
import openpyxl
import numpy as np
from scipy.stats import pearsonr

LEVEL_SCORE = {'A1': 1, 'A2': 2, 'B1': 3, 'B2': 4, 'C1': 5, 'C2': 6}

# 参照相关系数：v6 = phrase_cefr_tables.docx（旧抽取，固定级）；
#              fixed = 本项目 v12 抽取但用固定级（未 sense-aware）。
REF_R = {
    'PH_ratio_A1':      (-0.132, -0.109), 'PH_ratio_A2':  (None,   +0.007),
    'PH_ratio_B1':      (None,   +0.047), 'PH_ratio_B2':  (+0.042, +0.042),
    'PH_ratio_C1':      (+0.046, +0.046), 'PH_ratio_C2':  (None,   +0.009),
    'PH_ratio_A':       (-0.090, -0.096), 'PH_ratio_B':   (+0.066, +0.077),
    'PH_ratio_C':       (None,   +0.038), 'PH_ratio_AboveB1': (+0.090, +0.096),
    'PH_ratio_AboveB2': (+0.060, +0.060), 'PH_mean_score':(+0.100, +0.097),
    'PH_high_low_ratio':(+0.078, +0.081), 'PH_score_var': (None,   +0.007),
    'PH_pv_ratio':      (+0.074, +0.070), 'PH_multi_ratio':(-0.073, -0.020),
}

INDEX_ORDER = [
    'PH_ratio_A1','PH_ratio_A2','PH_ratio_B1','PH_ratio_B2','PH_ratio_C1','PH_ratio_C2',
    'PH_ratio_A','PH_ratio_B','PH_ratio_C','PH_ratio_AboveB1','PH_ratio_AboveB2',
    'PH_mean_score','PH_high_low_ratio','PH_score_var','PH_pv_ratio','PH_multi_ratio',
]


def load_sense_levels(ambig_path):
    """多义表 → {(词条, 释义序号str): 等级}。"""
    wb = openpyxl.load_workbook(ambig_path, read_only=True); ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    J = {x: i for i, x in enumerate(h)}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        term = str(r[J['词条']]).strip() if r[J['词条']] else ''
        no   = str(r[J['释义序号']]).strip() if r[J['释义序号']] is not None else ''
        lv   = str(r[J['等级']]).strip() if r[J['等级']] else ''
        if term and no:
            out[(term, no)] = lv
    return out


def _essay_indices(hits):
    """hits: [{'level'(可能None), 'pos', 'multi'}]。返回指标 dict 或 None。

    两套分母（有意为之，见模块 docstring“与 Hu et al. 的口径差异 · D2”）：
      · 等级分布类 PH_ratio_* / mean_score / score_var / high_low_ratio
        分母 = 有有效等级的命中数 `tot`（not_applied / 消歧失败的命中无法定级，
        不进这些指标）。
      · 短语属性类 PH_pv_ratio / PH_multi_ratio
        分母 = 全部命中 `n_all`。短语动词、多义是命中的固有属性，与能否定级无关；
        尤其未消歧成功的多义命中【本身就是多义】，必须计入 multi 分母，否则系统性漏计。
    另输出 PH_level_coverage = tot / n_all，记录被等级类指标排除的比例，供透明核查。
    """
    n_all = len(hits)
    lv = [h['level'] for h in hits if h['level'] in LEVEL_SCORE]
    tot = len(lv)
    if tot == 0:
        return None
    c = Counter(lv)
    n = lambda k: c.get(k, 0)
    r = lambda x: x / tot
    nA = n('A1') + n('A2'); nB = n('B1') + n('B2'); nC = n('C1') + n('C2')
    aboveB1 = nB + nC; aboveB2 = n('B2') + nC
    scores = [LEVEL_SCORE[x] for x in lv]
    return {
        'PH_ratio_A1': r(n('A1')), 'PH_ratio_A2': r(n('A2')),
        'PH_ratio_B1': r(n('B1')), 'PH_ratio_B2': r(n('B2')),
        'PH_ratio_C1': r(n('C1')), 'PH_ratio_C2': r(n('C2')),
        'PH_ratio_A': r(nA), 'PH_ratio_B': r(nB), 'PH_ratio_C': r(nC),
        'PH_ratio_AboveB1': r(aboveB1), 'PH_ratio_AboveB2': r(aboveB2),
        'PH_mean_score': float(np.mean(scores)),
        # 高低对比比 = AboveB1 短语数 / A 级短语数。无 A 级短语时该比值未定义 → NaN
        # （correlations() 按列成对剔除 NaN），避免退化成原始计数污染量纲。
        'PH_high_low_ratio': (aboveB1 / nA) if nA > 0 else float('nan'),
        'PH_score_var': float(np.var(scores)),
        'PH_pv_ratio': sum(1 for h in hits if h['pos'] == 'phrasal verb') / n_all,
        'PH_multi_ratio': sum(1 for h in hits if h['multi'] == '是') / n_all,
        'PH_level_coverage': tot / n_all,
    }


def _script_id(essay_id):
    """作文号 → 试卷号：剥掉末尾的 task 后缀（_A1/_A2/_A3…）。
    FCE 每份试卷含 2 篇作文、共享一个 1–40 总分——这是 Hu et al. (2025) 的分析单位。
    例：TR3*0100*2000*02_A1 与 …_A2 同属试卷 TR3*0100*2000*02。"""
    return str(essay_id).rsplit('_', 1)[0]


def compute(hit_detail_path, sense_levels, unit='answer'):
    """unit='answer'：逐篇作文聚合（N≈2470，同卷两篇共享总分→非独立，见 D4）。
       unit='script'：按试卷把两篇作文的命中【合并】后再聚合（N=1236，一卷一分，
                      匹配论文的分析单位），作为 D4 独立性问题的稳健性旁证。"""
    key = (lambda e: e) if unit == 'answer' else _script_id
    wb = openpyxl.load_workbook(hit_detail_path, read_only=True); ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    I = {x: i for i, x in enumerate(h)}
    by_essay = defaultdict(list); score = {}
    n_poly_no_level = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        gid = key(row[I['essay_id']])
        multi = str(row[I['多义']]).strip()
        if multi == '是':
            # 多义：用消歧义项在多义表里的真实等级
            sense = str(row[I['消歧义项']]).strip()
            level = sense_levels.get((str(row[I['短语']]).strip(), sense), '')
            if level not in LEVEL_SCORE:
                level = None; n_poly_no_level += 1
        else:
            level = str(row[I['CEFR']]).strip()   # 单义：固定等级
        by_essay[gid].append({'level': level, 'pos': str(row[I['词性']]).strip(), 'multi': multi})
        score[gid] = row[I['总分']]   # 同一 gid 内总分一致（answer 模式恒真；script 模式已核验两篇一致）
    rows = []
    for gid, hits in by_essay.items():
        idx = _essay_indices(hits)
        sc = score.get(gid)
        if idx is None or sc is None or float(sc) <= 0:
            continue
        # 'essay_id' 列在 script 模式下存放试卷号（分组键），列名沿用以免下游 schema 变动
        idx['essay_id'] = gid; idx['total_score'] = float(sc); idx['n_phrases'] = len(hits)
        rows.append(idx)
    print(f'  [{unit}] 分组数={len(rows)}；多义命中无有效等级(not_applied/解析失败)不计入等级指标: {n_poly_no_level}')
    return rows


def correlations(rows):
    """每个指标与总分的 Pearson r。含 NaN 的指标（如 high_low_ratio 在无 A 级短语
    的作文上未定义）按列【成对剔除】NaN 后再算，并回报实际参与的样本数 n。"""
    y_all = np.array([r['total_score'] for r in rows], dtype=float)
    out = []
    for name in INDEX_ORDER:
        x_all = np.array([r.get(name, np.nan) for r in rows], dtype=float)
        mask = ~np.isnan(x_all)
        x, y = x_all[mask], y_all[mask]
        n = int(mask.sum())
        if n < 3 or np.std(x) == 0:
            r_, p_ = float('nan'), float('nan')
        else:
            r_, p_ = pearsonr(x, y)
        out.append((name, float(np.nanmean(x_all)), float(np.nanstd(x_all)), n, r_, p_))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--hit-detail', default='结果_v12/hit_detail.xlsx')
    ap.add_argument('--ambig', default='phrase_extraction/data/多义短语_v10.xlsx')
    ap.add_argument('--unit', choices=['answer', 'script'], default='answer',
                    help='分析单位：answer=逐篇作文(N≈2470，同卷两篇共享总分)；'
                         'script=按试卷聚合两篇作文的命中(N=1236，一卷一分，匹配 Hu et al. 2025)')
    ap.add_argument('--out-indices', default=None)
    ap.add_argument('--out-corr', default=None)
    args = ap.parse_args()
    # 默认输出名随分析单位切换：script 模式落到 *_perscript，避免覆盖逐篇结果
    suffix = '' if args.unit == 'answer' else '_perscript'
    if args.out_indices is None:
        args.out_indices = f'结果_v12/phrase_indices_v12{suffix}.csv'
    if args.out_corr is None:
        args.out_corr = f'结果_v12/phrase_corr_v12{suffix}.csv'

    sense_levels = load_sense_levels(args.ambig)
    rows = compute(args.hit_detail, sense_levels, unit=args.unit)
    print(f'有效{"作文" if args.unit=="answer" else "试卷"}数（有等级命中且已评分）: {len(rows)}')

    import csv
    cols = ['essay_id', 'total_score', 'n_phrases', 'PH_level_coverage'] + INDEX_ORDER
    with open(args.out_indices, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
        for r in rows: w.writerow({k: r.get(k, '') for k in cols})
    print(f'[保存] 每篇指标 → {args.out_indices}')

    corr = correlations(rows)
    print(f'\n{"指标":22}{"Mean":>8}{"SD":>7}{"N":>6}{"senseAware_r":>13}{"p":>9}   {"固定级r":>8}{"docx_v6":>9}')
    print('-' * 88)
    with open(args.out_corr, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f); w.writerow(['index','mean','sd','n','r_senseaware','p','r_v12_fixed','r_docx_v6'])
        for name, mean, sd, n_, r_, p_ in corr:
            sig = '***' if p_ < .001 else '**' if p_ < .01 else '*' if p_ < .05 else ''
            v6, fixed = REF_R.get(name, (None, None))
            fs = f'{fixed:+.3f}' if fixed is not None else 'n.s.'
            v6s = f'{v6:+.3f}' if v6 is not None else 'n.s.'
            print(f'{name:22}{mean:8.3f}{sd:7.3f}{n_:6d}{r_:+13.3f}{sig:<3}{p_:8.3f}   {fs:>8}{v6s:>9}')
            w.writerow([name, f'{mean:.4f}', f'{sd:.4f}', n_, f'{r_:.4f}', f'{p_:.4g}',
                        f'{fixed:.3f}' if fixed is not None else '',
                        f'{v6:.3f}' if v6 is not None else ''])
    print(f'\n[保存] 相关性对比 → {args.out_corr}')


if __name__ == '__main__':
    main()
