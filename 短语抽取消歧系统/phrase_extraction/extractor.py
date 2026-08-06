#!/usr/bin/env python3
"""
EVP 短语抽取模块（独立封装版）
==============================
职责：作文 XML → 句子 → 归一化 → 正则匹配 → 抽取结果。
只做抽取，不含 LLM 消歧（消歧为独立模块 disambiguation）。

输出两张表：
  · 句子表   sentences.xlsx  每句一行
  · 抽取结果 hits.xlsx       每次命中一行（短语级，无消歧列）

公开 API：
  from extractor import extract, load_patterns, FixedPreprocessor
  results = extract(essay_paths, unambig_regex_xlsx, ambig_xlsx)
  # results: List[(Answer, List[MatchHit])]

CLI：
  python extractor.py --essays-dir <dir> \
      --unambig data/单义短语_v12_regex.xlsx --ambig data/多义短语_v10.xlsx \
      --out-sentences sentences.xlsx --out-hits hits.xlsx

依赖：openpyxl, nltk（wordnet/averaged_perceptron_tagger/punkt）
     + 同目录 evp_linguistic_data / evp_regex_generator
"""

import re
import argparse
import xml.etree.ElementTree as ET
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from evp_linguistic_data import (
    CONTRACTIONS      as _CONTRACTIONS,
    IRREG_VERB        as _IRREG_VERB,
    POSS_TO_YOUR      as _POSS_TO_YOUR,
    REFL_TO_YOURSELF  as _REFL_TO_YOURSELF,
    PRONOUN_TO_SB     as _PRONOUN_TO_SB,
)
from evp_regex_generator import generate_regex

from nltk.tokenize import TreebankWordTokenizer
from nltk.tag import pos_tag
from nltk.stem import WordNetLemmatizer
from nltk.corpus import wordnet as wn


# ══════════════════════════════════════════════════════════════════
# 配置：黑名单（正则过宽、无法可靠识别的短语，加载时跳过）
# ══════════════════════════════════════════════════════════════════
EXCLUDE_PHRASES = {
    'have sb do sth', 'have sth stolen/taken, etc.', 'get sb/sth to do sth',
    'be found', 'get sth painted/repaired, etc.', 'will have',
    'to go', 'a thing', 'the first', 'the people', 'the Right/right',
    'be friends (with sb)', 'go out (LIGHT/FIRE)', 'be up to sth',
    'smell of/like; smell delicious/horrible, etc.',
    'build (sth) up or build up (sth)', 'that is (to say)',
    'go in', 'the moment (that)', 'the poor', 'What ...for?', 'I know',
}


# ══════════════════════════════════════════════════════════════════
# §1  NLTK 预处理器（句子归一化）
# ══════════════════════════════════════════════════════════════════
_SB_PLACEHOLDER = '_sb_'
_PUNCT_CLEANUP  = re.compile(r'\s+([.,!?;:])')
_MULTI_SPACE    = re.compile(r'  +')


def _penn_to_wn(tag: str):
    if tag.startswith('V'): return wn.VERB
    if tag.startswith('N'): return wn.NOUN
    if tag.startswith('J'): return wn.ADJ
    if tag.startswith('R'): return wn.ADV
    return None


class FixedPreprocessor:
    """句子归一化：缩写展开 → 分词 → 词形还原 → 代词归一。"""

    def __init__(self):
        self._tok = TreebankWordTokenizer()
        self._lem = WordNetLemmatizer()
        sorted_keys = sorted(_CONTRACTIONS.keys(), key=len, reverse=True)
        self._contraction_re = re.compile(
            '|'.join(re.escape(k) for k in sorted_keys), re.IGNORECASE
        )

    def _pre_expand(self, text: str) -> str:
        text = text.replace('’', "'")
        return self._contraction_re.sub(
            lambda m: _CONTRACTIONS.get(m.group(0).lower(), m.group(0)), text
        )

    def normalize(self, text: str) -> str:
        pre       = self._pre_expand(text)
        raw_spans = list(self._tok.span_tokenize(pre))
        raw_toks  = [pre[s:e] for s, e in raw_spans]

        expanded: List[Tuple[str, int, int]] = []
        for token, (os_, oe) in zip(raw_toks, raw_spans):
            lower_norm = token.lower().replace('’', "'")
            if lower_norm in _CONTRACTIONS:
                for sub in _CONTRACTIONS[lower_norm].split():
                    expanded.append((sub, os_, oe))
            else:
                expanded.append((token, os_, oe))

        # 词性标注前统一小写：稳定全大写学习者文本的标注与词形还原
        tagged      = pos_tag([w.lower() for w, _, _ in expanded])
        tagged_list = list(tagged)
        norm_parts  = []

        for idx, ((word, _, _), (_, tag)) in enumerate(zip(expanded, tagged_list)):
            word_lower = word.lower()
            next_pos   = tagged_list[idx + 1][1] if idx + 1 < len(tagged_list) else ''
            wn_pos     = _penn_to_wn(tag)

            if wn_pos == wn.VERB:
                norm_word = _IRREG_VERB.get(word_lower) or self._lem.lemmatize(word_lower, wn.VERB)
            else:
                norm_word = word_lower

            if word_lower == "'s":
                if tag == 'VBZ':
                    norm_word = 'be'
                elif tag == 'VHZ':
                    norm_word = 'have'
                elif tag == 'POS' and next_pos in ('VBG', 'VBN', 'VBZ', 'JJ', 'JJR'):
                    norm_word = 'be'

            if norm_word == 'would' and next_pos == 'VBN':
                norm_word = 'have'

            if tag == 'PRP$' and norm_word in (_POSS_TO_YOUR | {'her'}):
                norm_word = 'your'
            elif tag != 'PRP$' and norm_word in _POSS_TO_YOUR:
                norm_word = 'your'

            if norm_word in _REFL_TO_YOURSELF:
                norm_word = 'yourself'
            elif norm_word in _PRONOUN_TO_SB:
                if not (norm_word == 'her' and tag != 'PRP'):
                    norm_word = _SB_PLACEHOLDER

            norm_parts.append(norm_word)

        normed = _PUNCT_CLEANUP.sub(r'\1', ' '.join(norm_parts))
        return _MULTI_SPACE.sub(' ', normed).strip()


# ══════════════════════════════════════════════════════════════════
# §2  XML 解析：提取原始学习者文本 + NS 错误标注
# ══════════════════════════════════════════════════════════════════
@dataclass
class Answer:
    essay_id:    str
    question_no: str
    exam_score:  Optional[float]
    total_score: Optional[float]
    paragraphs:  List[str]
    error_types: List[str]
    para_ns_spans: List[List[Tuple[int, str]]] = field(default_factory=list)


def _extract_raw_text(elem) -> Tuple[str, List[Tuple[int, str]]]:
    """返回 (归一化文本, [(错误在文本中的字符偏移, NS错误类型)])。

    在每个 NS 更正内容的起点插入零宽标记 \\x00，归一化后据标记位置算偏移、
    再移除标记（并吞掉标记右侧多余空格），使文本与不插标记时逐字符一致。
    偏移用于把每个错误归属到所在句子。"""
    MARK = '\x00'
    parts, ns_types = [], []

    def _walk(node, inside_i=False):
        if node.text:
            parts.append(node.text)
        for child in node:
            tag = child.tag
            if tag == 'NS':
                ns_type = child.get('type', '')
                if ns_type:
                    ns_types.append(ns_type)
                    parts.append(MARK)
                i_elem = child.find('i')
                if i_elem is not None:
                    _walk(i_elem, inside_i=True)
            else:
                if tag == 'i':
                    _walk(child, inside_i=True)
                elif tag == 'c':
                    pass
                else:
                    _walk(child, inside_i=inside_i)
            if child.tail:
                parts.append(child.tail)

    _walk(elem)
    marked = re.sub(r'\s+', ' ', ''.join(parts)).strip()
    chars, spans, ti, skip_space = [], [], 0, False
    for ch in marked:
        if ch == MARK:
            if ti < len(ns_types):
                spans.append((len(chars), ns_types[ti])); ti += 1
            skip_space = bool(chars) and chars[-1] == ' '
            continue
        if ch == ' ' and skip_space:
            skip_space = False
            continue
        skip_space = False
        chars.append(ch)
    clean = ''.join(chars)
    # 标记落在文本首尾时可能残留边缘空格；再 strip 并按左移量平移偏移，
    # 使 clean 与不插标记的文本逐字符一致。
    lstripped = clean.lstrip()
    shift = len(clean) - len(lstripped)
    clean = lstripped.rstrip()
    if clean:
        hi = len(clean) - 1   # 尾部错误(如句末缺标点)夹到最后一个字符，归入末句
        spans = [(min(max(0, off - shift), hi), t) for off, t in spans]
    else:
        spans = []
    return clean, spans


def parse_clc_xml(xml_path: str) -> List[Answer]:
    """解析一个 CLC XML 文件，返回 Answer 列表（每个 answer 一个）。"""
    tree    = ET.parse(xml_path)
    root    = tree.getroot()
    head    = root if root.tag == 'head' else root.find('.//head')
    sortkey = head.get('sortkey', Path(xml_path).stem) if head is not None else Path(xml_path).stem

    score_elem  = root.find('.//score')
    def _to_float(s):
        try: return float(str(s).strip().rstrip('TtSs'))
        except (ValueError, TypeError): return None
    total_score = _to_float(score_elem.text) if score_elem is not None else None

    answers = []
    for ans_elem in root.findall('.//answer1') + root.findall('.//answer2') + \
                    root.findall('.//answer3') + root.findall('.//answer4'):
        q_no_elem    = ans_elem.find('question_number')
        score_e      = ans_elem.find('exam_score')
        question_no  = q_no_elem.text.strip() if q_no_elem is not None and q_no_elem.text else ans_elem.tag[-1]
        exam_score   = _to_float(score_e.text) if score_e is not None else None

        coded = ans_elem.find('coded_answer')
        if coded is None:
            continue

        paragraphs, para_spans, all_ns = [], [], []
        for p in coded.findall('p'):
            raw_text, ns_spans = _extract_raw_text(p)
            if raw_text:
                paragraphs.append(raw_text)
                para_spans.append(ns_spans)
            all_ns.extend(t for _, t in ns_spans)

        answers.append(Answer(
            essay_id    = f'{sortkey}_A{question_no}',
            question_no = question_no,
            exam_score  = exam_score,
            total_score = total_score,
            paragraphs  = paragraphs,
            error_types = sorted(set(all_ns)),
            para_ns_spans = para_spans,
        ))
    return answers


# ══════════════════════════════════════════════════════════════════
# §3  正则模式加载
# ══════════════════════════════════════════════════════════════════
@dataclass
class PhrasePattern:
    phrase_id:    str
    phrase_form:  str
    cefr:         str
    pos:          str
    meaning:      str
    regex:        str
    compiled:     re.Pattern
    is_ambiguous: bool
    senses:       List[Dict]


def load_patterns(unambig_path: str, ambig_path: str) -> List[PhrasePattern]:
    """加载单义(含正则) + 多义短语，合并为 PhrasePattern 列表，标注 is_ambiguous。"""
    ambig_senses: Dict[str, List[Dict]] = defaultdict(list)
    wb_a = openpyxl.load_workbook(ambig_path, read_only=True)
    ws_a = wb_a.active
    ah   = [c.value for c in next(ws_a.iter_rows(min_row=1, max_row=1))]
    ac   = {v: i for i, v in enumerate(ah)}
    for row in ws_a.iter_rows(min_row=2, values_only=True):
        phrase = str(row[ac['词条']]).strip() if row[ac['词条']] else ''
        if not phrase:
            continue
        ambig_senses[phrase].append({
            '释义序号': row[ac.get('释义序号', ac.get('义项序号', 0))],
            '释义':     row[ac.get('释义', 0)] or '',
            '词典例句': row[ac.get('词典例句', 0)] or '',
            '等级':     row[ac.get('等级', 0)] or '',
        })
    wb_a.close()
    ambig_set = set(ambig_senses.keys())
    print(f'  多义短语词条数：{len(ambig_set)}')

    wb_u = openpyxl.load_workbook(unambig_path, read_only=True)
    ws_u = wb_u.active
    uh   = [c.value for c in next(ws_u.iter_rows(min_row=1, max_row=1))]
    uc   = {v: i for i, v in enumerate(uh)}
    missing = [c for c in ['原始短语','拆分后','CEFR等级','词性','释义','正则表达式'] if c not in uc]
    if missing:
        print(f'  [警告] 单义表缺少列名：{missing} | 实际：{[k for k in uc if k is not None]}')

    patterns, n_ok, n_err, n_excluded = [], 0, 0, 0
    for row in ws_u.iter_rows(min_row=2, values_only=True):
        orig    = str(row[uc['原始短语']]).strip()   if '原始短语'   in uc and row[uc['原始短语']]   else ''
        phrase  = str(row[uc['拆分后']]).strip()     if '拆分后'     in uc and row[uc['拆分后']]     else ''
        cefr    = str(row[uc['CEFR等级']] or '').strip() if 'CEFR等级'   in uc else ''
        pos     = str(row[uc['词性']]     or '').strip() if '词性'       in uc else ''
        meaning = str(row[uc['释义']]     or '').strip() if '释义'       in uc else ''
        regex   = str(row[uc['正则表达式']] or '').strip() if '正则表达式' in uc else ''

        if not phrase or not regex or regex in ('None', 'nan', ''):
            continue
        if orig in EXCLUDE_PHRASES:
            n_excluded += 1; continue
        if len(phrase.rstrip('-').split()) == 1:
            n_excluded += 1; continue
        if orig == '(the) most attractive/important/popular, etc.' and not phrase.startswith('the '):
            n_excluded += 1; continue
        try:
            compiled = re.compile(regex)
        except re.error:
            n_err += 1; continue

        is_ambig = orig in ambig_set
        patterns.append(PhrasePattern(
            phrase_id=orig, phrase_form=phrase, cefr=cefr, pos=pos, meaning=meaning,
            regex=regex, compiled=compiled, is_ambiguous=is_ambig,
            senses=ambig_senses[orig] if is_ambig else [],
        ))
        n_ok += 1
    wb_u.close()

    # 多义补充：多义表有、单义表无的词条，直接用词条生成正则
    covered = set(p.phrase_id for p in patterns)
    n_ambig_added = n_ambig_err = 0
    for phrase_id, senses in ambig_senses.items():
        if phrase_id in covered or phrase_id in EXCLUDE_PHRASES or len(phrase_id.split()) == 1:
            continue
        try:
            regex_str, _ = generate_regex(phrase_id)
            if not regex_str:
                continue
            compiled = re.compile(regex_str)
        except re.error:
            n_ambig_err += 1; continue
        rep = senses[0]
        patterns.append(PhrasePattern(
            phrase_id=phrase_id, phrase_form=phrase_id, cefr=str(rep.get('等级','') or ''),
            pos='phrase', meaning=str(rep.get('释义','') or ''), regex=regex_str,
            compiled=compiled, is_ambiguous=True, senses=senses,
        ))
        n_ambig_added += 1

    print(f'  正则加载：{n_ok} 成功，{n_excluded} 黑名单跳过，{n_err} 正则错误')
    print(f'  多义补充：{n_ambig_added} 条，{n_ambig_err} 错误')
    return patterns


# ══════════════════════════════════════════════════════════════════
# §4  正则匹配 + 最长匹配去重 + 词性验证
# ══════════════════════════════════════════════════════════════════
@dataclass
class MatchHit:
    pattern:    PhrasePattern
    sentence:   str
    norm:       str
    match_span: Tuple[int, int]
    sent_idx:   int
    context:    str
    sent_ns:    List[str] = field(default_factory=list)


def _sentences_from_answer(answer: Answer) -> List[tuple]:
    """按句切分，并把落在该句字符区间内的 NS 错误归属到该句。"""
    sents = []
    spans_per_para = (answer.para_ns_spans if answer.para_ns_spans
                      else [[] for _ in answer.paragraphs])
    for para, spans in zip(answer.paragraphs, spans_per_para):
        # 句子边界：句末标点(.!?)后接空白；用字符区间覆盖整段
        bounds = [m.end() for m in re.finditer(r'[.!?]+\s+', para)]
        starts, ends = [0] + bounds, bounds + [len(para)]
        for st, en in zip(starts, ends):
            s = para[st:en].strip()
            if not s:
                continue
            sent_ns = sorted({t for off, t in spans if st <= off < en})
            sents.append((s, sent_ns))
    return sents


def _longest_match(hits: List[MatchHit]) -> List[MatchHit]:
    if not hits:
        return hits
    sorted_hits = sorted(hits, key=lambda h: h.match_span[1] - h.match_span[0], reverse=True)
    kept = []
    for h in sorted_hits:
        hs, he = h.match_span
        dominated = any(ks <= hs and he <= ke for k in kept for ks, ke in [k.match_span])
        if not dominated:
            kept.append(h)
    return kept


# ── §4b 词性验证 ────────────────────────────────────────────────────
_ANCHOR_SKIP = {'or','and','but','be','been','being','my','your','his','her',
                'our','their','these','those','each','every'}
_PARTICLES = {'back','up','down','off','out','on','in','over','away','around','through'}


def _extract_anchor_sequence(phrase_id: str) -> List[str]:
    p = phrase_id.lower()
    p = re.sub(r'\(.*?\)', '', p)
    p = re.sub(r',?\s*etc\.?', '', p)
    p = re.sub(r"\bsb(?:'s)?\b|\bsth\b|\bswh\b|\bdoing\b", ' \x00 ', p, flags=re.IGNORECASE)
    p = re.sub(r'/\w+', '', p)
    p = re.sub(r'[^\w\s\x00]', ' ', p)
    return [w for w in p.split() if w != '\x00' and len(w) >= 1 and w not in _ANCHOR_SKIP]


def _is_sandwiched(anchor_seq: List[str]) -> bool:
    return bool(anchor_seq) and anchor_seq[-1] in _PARTICLES


def _check_noun_after_anchor(sentence: str, phrase_id: str) -> bool:
    from nltk.tokenize import TreebankWordTokenizer as _TWT
    from nltk.tag import pos_tag as _pos_tag
    from nltk.stem import WordNetLemmatizer as _WNL
    from nltk.corpus import wordnet as _wn

    anchor_seq = _extract_anchor_sequence(phrase_id)
    if not anchor_seq:
        return True

    _lem = _WNL()
    def _lemv(w):
        w = w.lower()
        return _IRREG_VERB.get(w) or _lem.lemmatize(w, _wn.VERB)

    tokens  = _TWT().tokenize(sentence)
    tagged  = _pos_tag(tokens)
    t_lemma = [_lemv(t) for t in tokens]
    a_lemma = [_lemv(a) for a in anchor_seq]

    NOUN_TAGS    = {'NN','NNS','NNP','NNPS','PRP','WP'}
    REJECT_FIRST = {'IN','TO','RB','RBR','RBS'}
    SKIP_DET     = {'the','a','an','not','also','just','only','really','very','quite'}

    target = a_lemma[0] if _is_sandwiched(anchor_seq) else a_lemma[-1]
    anchor_pos = -1
    for i, tl in enumerate(t_lemma):
        if tl == target:
            anchor_pos = i
    if anchor_pos == -1:
        return True

    window = tagged[anchor_pos + 1: anchor_pos + 6]
    if not window:
        return False
    for word, tag in window:
        if word.lower() in SKIP_DET:
            continue
        if tag in REJECT_FIRST:
            return False
        break
    return any(tag in NOUN_TAGS for _, tag in window)


def _verify_make_sb_do(sentence: str) -> bool:
    from nltk.tokenize import TreebankWordTokenizer as _TWT
    from nltk.tag import pos_tag as _pos_tag
    tokens = _TWT().tokenize(sentence)
    tagged = _pos_tag(tokens)
    for i, (w, t) in enumerate(tagged):
        if w.lower() in ('make', 'makes', 'made', 'making'):
            if i + 1 >= len(tagged):
                return False
            next_w, next_t = tagged[i + 1]
            if next_t in ('DT', 'PDT'):
                return False
            if next_t in ('PRP', 'NN', 'NNS', 'NNP'):
                for j in range(i + 2, min(i + 5, len(tagged))):
                    if tagged[j][1] in ('VB','VBP','VBZ','VBD','VBG','VBN'):
                        return True
                return True
            if next_t == 'TO':
                return True
    return False


POS_VALIDATORS: Dict[str, callable] = {
    'make sb do sth': _verify_make_sb_do,
}


def pos_validate(hit: 'MatchHit') -> bool:
    validator = POS_VALIDATORS.get(hit.pattern.phrase_id)
    if validator is not None:
        try:
            return validator(hit.sentence)
        except Exception:
            return True
    pid = hit.pattern.phrase_id.lower()
    if re.search(r"\bsb(?:'s)?\b|\bsth\b", pid):
        try:
            return _check_noun_after_anchor(hit.sentence, hit.pattern.phrase_id)
        except Exception:
            return True
    return True


def match_answer(answer: Answer,
                 patterns: List[PhrasePattern],
                 preprocessor: FixedPreprocessor) -> List[MatchHit]:
    """对一篇 answer 的所有句子跑正则，返回去重后的命中列表。"""
    sents    = _sentences_from_answer(answer)
    all_hits = []
    for sent_idx, (sent, sent_ns) in enumerate(sents):
        try:
            norm = preprocessor.normalize(sent)
        except Exception:
            norm = sent.lower()

        ctx_parts = []
        if sent_idx > 0:
            ctx_parts.append(sents[sent_idx - 1][0])
        ctx_parts.append(sent)
        if sent_idx < len(sents) - 1:
            ctx_parts.append(sents[sent_idx + 1][0])
        context = ' '.join(ctx_parts)

        sent_hits: List[MatchHit] = []
        for pat in patterns:
            for m in pat.compiled.finditer(norm):
                sent_hits.append(MatchHit(
                    pattern=pat, sentence=sent, norm=norm,
                    match_span=(m.start(), m.end()), sent_idx=sent_idx,
                    context=context, sent_ns=sent_ns,
                ))

        sent_hits = _longest_match(sent_hits)
        sent_hits = [h for h in sent_hits if pos_validate(h)]

        seen_in_sent, deduped = set(), []
        for h in sent_hits:
            if h.pattern.phrase_id not in seen_in_sent:
                seen_in_sent.add(h.pattern.phrase_id)
                deduped.append(h)
        all_hits.extend(deduped)
    return all_hits


# ══════════════════════════════════════════════════════════════════
# §5  输出
# ══════════════════════════════════════════════════════════════════
def extract(essay_paths: List[str], unambig_regex_xlsx: str, ambig_xlsx: str
            ) -> List[Tuple[Answer, List[MatchHit]]]:
    """
    抽取主入口。返回 [(Answer, [MatchHit]), ...]。
    unambig_regex_xlsx: 含"正则表达式"列的单义表（如 data/单义短语_v12_regex.xlsx）。
    ambig_xlsx: 多义短语义项表。
    """
    patterns = load_patterns(unambig_regex_xlsx, ambig_xlsx)
    pre = FixedPreprocessor()
    results = []
    for path in essay_paths:
        try:
            answers = parse_clc_xml(path)
        except Exception as e:
            print(f'  [跳过] 解析失败 {path}: {e}')
            continue
        for a in answers:
            results.append((a, match_answer(a, patterns, pre)))
    return results


def write_sentences(results, out_path: str):
    """句子表：每句一行。"""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '句子表'
    hdr = ['essay_id', '题号', '总分', '句序', '句子', '段落错误类型']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDEEFF')
    for answer, _ in results:
        for idx, (sent, ns) in enumerate(_sentences_from_answer(answer)):
            ws.append([answer.essay_id, answer.question_no, answer.total_score,
                       idx, sent, ','.join(ns)])
    for i, w in enumerate([22, 6, 6, 6, 90, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row: c.alignment = Alignment(vertical='top', wrap_text=True)
    wb.save(out_path)
    print(f'[保存] 句子表 → {out_path}')


def write_hits(results, out_path: str):
    """抽取结果：每次命中一行（短语级，无消歧列）。"""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '抽取结果'
    hdr = ['essay_id', '题号', '句序', '句子', '短语', '变体形式',
           'CEFR', '词性', '释义', '多义']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDEEFF')
    for answer, hits in results:
        for h in hits:
            ws.append([answer.essay_id, answer.question_no, h.sent_idx, h.sentence,
                       h.pattern.phrase_id, h.pattern.phrase_form, h.pattern.cefr,
                       h.pattern.pos, h.pattern.meaning,
                       '是' if h.pattern.is_ambiguous else '否'])
    for i, w in enumerate([22, 6, 6, 60, 26, 24, 8, 8, 40, 6], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row: c.alignment = Alignment(vertical='top', wrap_text=True)
    wb.save(out_path)
    print(f'[保存] 抽取结果 → {out_path}')


def main():
    _here = Path(__file__).parent
    ap = argparse.ArgumentParser(description='EVP 短语抽取（独立模块，无消歧）')
    ap.add_argument('--essays', nargs='+', help='XML 文件路径列表')
    ap.add_argument('--essays-dir', help='作文 XML 目录（递归 *.xml）')
    ap.add_argument('--unambig', default=str(_here / 'data' / '单义短语_v12_regex.xlsx'),
                    help='含正则的单义短语表')
    ap.add_argument('--ambig', default=str(_here / 'data' / '多义短语_v10.xlsx'),
                    help='多义短语义项表')
    ap.add_argument('--out-sentences', default='sentences.xlsx')
    ap.add_argument('--out-hits', default='hits.xlsx')
    args = ap.parse_args()

    paths = list(args.essays or [])
    if args.essays_dir:
        paths += [str(p) for p in Path(args.essays_dir).rglob('*.xml')]
    if not paths:
        ap.error('请指定 --essays 或 --essays-dir')

    print(f'[INFO] 作文文件：{len(paths)}')
    results = extract(paths, args.unambig, args.ambig)
    n_ans = len(results); n_hits = sum(len(h) for _, h in results)
    print(f'[INFO] answer 数：{n_ans} | 命中总数：{n_hits}')
    write_sentences(results, args.out_sentences)
    write_hits(results, args.out_hits)


if __name__ == '__main__':
    main()
