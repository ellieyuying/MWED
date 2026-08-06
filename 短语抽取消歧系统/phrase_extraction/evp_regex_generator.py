#!/usr/bin/env python3
"""
EVP 短语 → 正则表达式生成器。

phrase_to_regex 把标准化短语转成正则，配合预处理器在归一化文本上匹配：
预处理器已完成词形还原、代词/物主归一、缩写展开、小写化，故正则只写词原形
与归一化占位符（sb→_sb_、your、be 等）。

对外接口：generate_regex / phrase_to_regex（短语→正则）；process_file（批量写入正则列）。
短语表仅含多词短语，不含单词条目。
"""

import openpyxl
import re
from collections import defaultdict
from evp_linguistic_data import (
    VERB_FORMS, INFLECTION_TO_BASE,
    CONTRACTIONS as _PHRASE_CONTRACTIONS,
    IRREG_VERB   as _IRREG_VERB,
    _TIMEUNIT_RE, _NUM_RE, _DET, _SBPOS_RE,
    JJ_ING_ADJECTIVES as _JJ_ING_ADJECTIVES,
    JJ_ED_ADJECTIVES  as _JJ_ED_ADJECTIVES,
    PHRASE_VERB_SKIP  as _PHRASE_VERB_SKIP,
    VERB_NEG_LA       as _VERB_NEG,
    SB_NOT_VERB_LA    as _SB_NOT_VERB,
)
from evp_special_cases import match_special_case, MERGE_PATTERNS

# 动词原形还原：与预处理器 FixedPreprocessor 一致（不规则表 + WordNet 动词还原），
# 使短语侧生成的原形与句子归一化后的形态对齐。
_LEM = None
def _verb_lemma(w):
    global _LEM
    if _LEM is None:
        from nltk.stem import WordNetLemmatizer
        _LEM = WordNetLemmatizer()
    return _IRREG_VERB.get(w) or _LEM.lemmatize(w, 'v')

# ─────────────────────────────────────────────────────────────
# 2. 占位符
# ─────────────────────────────────────────────────────────────
PH_ELLIPSIS  = '\x00ELL\x00'
PH_NUM       = '\x00NUM\x00'
PH_SBPOS     = '\x00SBPOS\x00'
PH_SB        = '\x00SB\x00'
PH_STH       = '\x00STH\x00'
PH_SWH       = '\x00SWH\x00'    # swh (somewhere) 占位符
PH_DO_STH    = '\x00DOSTH\x00'    # do sth 元占位符
PH_DOING_STH = '\x00DOINGSTH\x00' # doing sth 元占位符
# 泛化占位符（来自输入表的 _adj_/_noun_/_adv_/_verb_/_timeunit_ 标记）
PH_ADJ      = '\x00ADJ\x00'      # _adj_：形容词/过去分词位置
PH_NOUN     = '\x00NOUN\x00'     # _noun_：普通名词位置
PH_ADV      = '\x00ADV\x00'      # _adv_：副词位置
PH_VERB     = '\x00VERB\x00'     # _verb_：动词位置
PH_TIMEUNIT = '\x00TIMEUNIT\x00' # _timeunit_：时间单位位置

# ── do/doing sth 占位符展开式 ─────────────────────────────────
# _VERB_NEG 为功能词负向前瞻，拼接词数量词组成完整展开式。
# do sth：原形动词后跟 0-4 个词
_PH_DO_STH_RE    = _VERB_NEG + r'\w+(?:\s+\S+){0,4}'
# doing sth：归一化后 gerund 已还原为原形，同样用 \w+ 匹配原形动词
_PH_DOING_STH_RE = _VERB_NEG + r'\w+(?:\s+\S+){0,3}'

def _expand_contractions(phrase: str) -> str:
    """展开短语中的缩写，使生成的正则与预处理后的文本一致。"""
    def replace(m):
        token = m.group(0).lower().replace('\u2019', "'")
        return _PHRASE_CONTRACTIONS.get(token, m.group(0))
    return re.sub(r"[A-Za-z]+(?:['\u2019][A-Za-z]+)+", replace, phrase, flags=re.I)


def escape_for_regex(s):
    result = []
    for ch in s:
        if ch in r'\^$*+?{}()|[]':
            result.append('\\' + ch)
        else:
            result.append(ch)
    return ''.join(result)

# ─────────────────────────────────────────────────────────────
# 6. 核心正则生成
# ─────────────────────────────────────────────────────────────
def phrase_to_regex(phrase):
    """
    将标准化短语转换为正则表达式。

    预处理层已完成：词形还原（made→make）、物主代词归一（his/her/my→your）、
    人称代词替换（he/she/they→_sb_）、缩写展开（don't→do not）、小写化。
    因此正则生成器只需写词的原形，无需展开变形。

    返回 (regex_str, notes_str)
    """
    # ── 特例规则表（见 evp_special_cases）：命中即返回固定/半固定正则 ──
    _special = match_special_case(phrase)
    if _special is not None:
        return _special

    if not phrase or not str(phrase).strip():
        return '', ''

    p = str(phrase).strip()
    notes = []

    # ── Pre-step: 括号可选处理 ────────────────────────────────
    # EVP 记法中 (X) 表示 X 可选（如 "keep up (with sb/sth)"）。此处对残留括号做含/不含两版展开，
    # 递归各自生成正则再 OR；(?i) 提到最外层，避免内嵌 (?:...(?i)...) 报错。
    # 放在特例块之后：that is (to say) / or something (like that) 等已由特例返回。
    if re.search(r'\([^)]*\)', p):
        _with    = re.sub(r'\s+', ' ', re.sub(r'[()]', '', p)).strip()          # 去括号留内容
        _without = re.sub(r'\s+', ' ', re.sub(r'\s*\([^)]*\)\s*', ' ', p)).strip()  # 去括号连内容
        _rgx_w, _note_w = phrase_to_regex(_with)
        if _without and _without != _with and _rgx_w:
            _rgx_wo, _ = phrase_to_regex(_without)
            _inner_w  = re.sub(r'^\(\?i\)', '', _rgx_w)
            _inner_wo = re.sub(r'^\(\?i\)', '', _rgx_wo)
            if _rgx_wo:
                return '(?i)(?:' + _inner_w + '|' + _inner_wo + ')', (_note_w + '|括号可选')
        return _rgx_w, _note_w

    # ── Pre-step: 短语端缩写展开（与预处理器保持一致）──────────
    p_expanded = _expand_contractions(p)
    if p_expanded != p:
        notes.append('缩写展开')
        p = p_expanded

    # ── Pre-step 0.5-pre: 固定搭配词形还原前保护 ───────────────────
    # used to：还原会把 used→use；be going to：going→go 后 Step 0b 检测不到
    _has_used_to = bool(re.search(r'\bused\s+to\b', p, re.IGNORECASE))
    if _has_used_to:
        p = re.sub(r'\bused\s+to\b', '\x00USEDTO\x00', p, flags=re.IGNORECASE)
    _has_be_going_to = bool(re.match(r'^be\s+going\s+to\b', p, re.IGNORECASE))
    if _has_be_going_to:
        p = re.sub(r'^be\s+going\s+to\b', '\x00BEGOINGTO\x00', p, flags=re.IGNORECASE)
        notes.append('be going to→固定前缀')

    # ── Pre-step 0.5a: cannot → can not ──────────────────────────
    # matcher 把 cannot 分词为 ['can', 'not'] 两个 token，
    # norm 里是 "can not"；正则若写 cannot 则无法匹配。
    p = re.sub(r'\bcannot\b', 'can not', p, flags=re.IGNORECASE)

    # ── Pre-step 0.5: 短语动词词形还原─────────────────
    # 规则还原：按优先级依次判断，不依赖词性标注器。
    # 优先级：跳过词 > -ed 黑名单（备选正则）> -ing 黑名单（保留）
    #         > INFLECTION_TO_BASE 查表 > 后缀规则推断 > 保留原形
    _new_toks = []
    for _tok in p.split():
        _tl = _tok.lower()
        if '\x00' in _tok or _tl in _PHRASE_VERB_SKIP:
            _new_toks.append(_tok)
        elif _tl in _JJ_ING_ADJECTIVES:
            _new_toks.append(_tok)               # -ing 形容词：保留
        elif (_tl.endswith('ed') or _tl in _JJ_ED_ADJECTIVES
              or _tl in INFLECTION_TO_BASE):
            # 分词/过去式/不规则形（-ed、broken/torn 等）：预处理器按语境把它还原为
            # 动词原形或保留原形态，故生成 (?:原形态|原形) 两存、与预处理器输出对齐。
            # 不规则形优先查 INFLECTION_TO_BASE，其次 WordNet。
            _lemma = INFLECTION_TO_BASE.get(_tl) or _verb_lemma(_tl)
            if _lemma and _lemma != _tl:
                _new_toks.append(f'\x00EDALTS\x00{_tl}|{_lemma}\x00ENDALT\x00')
            else:
                _new_toks.append(_tok)
        elif _tl.endswith('ing') and _tl[:-3] in VERB_FORMS:
            _new_toks.append(_tl[:-3])           # e.g. making → make（仅限已知动词）
        elif _tl.endswith('ing') and _tl[:-4] in VERB_FORMS:
            _new_toks.append(_tl[:-4])           # e.g. coming → come（去 e 后加 ing）
        elif _tl.endswith('s') and not _tl.endswith('ss') and _tl[:-1] in VERB_FORMS:
            _new_toks.append(_tl[:-1])           # e.g. makes → make
        elif _tl in ('can', 'may'):
            # 情态词同族替换：can↔could、may↔might（例句常用过去式）
            _alt = 'can|could' if _tl == 'can' else 'may|might'
            _new_toks.append(f'\x00EDALTS\x00{_alt}\x00ENDALT\x00')
        else:
            _new_toks.append(_tok)
    _p_lemmed = ' '.join(_new_toks)
    if _p_lemmed != p:
        notes.append('动词还原')
        p = _p_lemmed

    # ── Pre-step 2: 短语端代词归一（与预处理器保持一致）──────────
    # A. 反身代词归一 → yourself
    p = re.sub(r'\b(himself|herself|themselves|myself|ourselves|oneself)\b',
               'yourself', p, flags=re.IGNORECASE)
    # B. 物主代词归一 → your（含 one's）
    p = re.sub(r"\bone's\b", 'your', p, flags=re.IGNORECASE)
    p = re.sub(r'\b(their|my|his|our)\b', 'your', p, flags=re.IGNORECASE)
    # C. 宾格代词归一 → sb
    p = re.sub(r'\b(me|him|her|them|us)\b', 'sb', p, flags=re.IGNORECASE)
    # D：somewhere / something → 通配占位符
    #   somewhere → swh（已有 1-2 词通配逻辑）
    #   something → sth（已有 1-3 词通配逻辑）
    p = re.sub(r'\bsomewhere\b', 'swh', p, flags=re.IGNORECASE)
    # something 字面词保护：开头 or 前面是 or/of/be 时视为字面词，不替换为 sth
    if re.search(r'\bsomething\b', p, re.IGNORECASE):
        _is_literal_something = bool(
            re.match(r'^something\b', p, re.IGNORECASE) or
            re.search(r'(?:\bor\b|\bof\b|\bbe\b)\s+something\b', p, re.IGNORECASE)
        )
        if not _is_literal_something:
            p = re.sub(r'\bsomething\b', 'sth', p, flags=re.IGNORECASE)

    # ── Pre-step 3: 撇号拆分（word's → word 's），跳过 sb's ────
    # sb's 整体由 Step 2 处理，不预拆；其他 word's 正常拆分
    p = re.sub(r"(?<!\bsb)'s\b", r" 's", p)

    # ── Pre-step 4: had/have better → (?:have|would) better ─────
    # normalize_phrase 会将 had→have（VBD 词形还原），
    # 因此必须同时捕获 had better 和 have better 两种形式。
    if re.search(r'\b(?:had|have)\s+better\b', p, re.IGNORECASE):
        p = re.sub(r'\b(?:had|have)\s+better\b', '\x00HADBETTER\x00', p, flags=re.IGNORECASE)
        notes.append('had better→(?:have|would)')

    # ── Step 0: decade合并 ────────────────────────────────────
    for pattern, merged_regex, merge_note in MERGE_PATTERNS:
        if pattern.match(p):
            return merged_regex, merge_note

    # ── Step 0b: 开头 be → 谓语前近邻必需 ──────────────────────
    # be 常被主语/副词与谓语分隔；若整词删去，正则只剩裸谓语会到处误匹配
    # （be around→任意 around、be called→主动 call）。故不删 be，改在 Step 9
    # 组装时要求谓语前近邻出现系动词，间隔 ≤ BE_GAP 词以容纳 "be always around"。
    # be going to 已在 Pre-step 0.5-pre 转为 BEGOINGTO 占位符，不会到达这里。
    _leading_be = False
    if re.match(r'^be\s+', p, re.IGNORECASE):
        p = re.sub(r'^be\s+', '', p, flags=re.IGNORECASE)
        _leading_be = True
        notes.append('开头be→近邻必需')

    # ── Step 0c: 词汇并列·etc. 处理 ──────────────────────────────
    # 短语中 word1/word2[/word3], etc. 的斜杠列举词展开为 (?:词1|词2|词3) 枚举正则。
    # 斜杠无空格 = 左右选其一，与固定词共同构成变体；若用 \w+ 通配会匹配几乎所有
    # 含该动词的句子（如 "go bald/blind/grey, etc."），产生大量误识别。
    # 斜杠两侧有空格的情况（如 "a pile of / piles of sth"）在输入表阶段已拆分，
    # 不会以未拆分形式到达这里。
    if 'etc.' in p.lower():
        # 枚举展开：把斜杠列举组替换为带边界标记的枚举串，
        # 边界标记在 Step 7 后统一还原为正则 (?:词1|词2|词3)，
        # 避免中间步骤的 escape_for_regex 破坏括号/管道符。
        _ENUM_L = '\x01EL\x01'   # 枚举组左边界
        _ENUM_R = '\x01ER\x01'   # 枚举组右边界
        _ENUM_P = '\x01EP\x01'   # 枚举词分隔符
        _ENUM_WIDE = '\x01EW\x01'  # 宽槽：还原为 \w+（列举后有固定名词锚定时用）
        def _expand_slash_mark(m):
            words = [w.lower() for w in m.group(0).split('/') if w]
            if len(words) == 1:
                return words[0]
            # etc. 表"及类似词"：若列举组【后面还有真正的固定内容词】，说明有锚定，
            # 放宽为 \w+ 覆盖未列出的同类词（do a good/excellent,etc.job → 也匹配
            # "a great job"）。占位符 sth/sb/swh 是通配、不算锚（否则 run sth
            # along/over,etc.sth → "run 任意词" 过匹配）；无固定锚则保持枚举。
            rest = re.sub(r',?\s*etc\.|\b(?:sth|sb|swh)\b', '', m.string[m.end():],
                          flags=re.IGNORECASE)
            if re.search(r'[a-z]{2,}', rest, re.IGNORECASE):
                return _ENUM_WIDE
            return _ENUM_L + _ENUM_P.join(words) + _ENUM_R
        # 扩展匹配模式：同时捕获含连字符的列举词（如 hair/make-up）
        p = re.sub(r'\b\w+(?:[/-]\w+)+\b', _expand_slash_mark, p, flags=re.IGNORECASE)
        p = re.sub(r',?\s*etc\.', '', p, flags=re.IGNORECASE)
        p = p.strip()
        notes.append('etc.列举→枚举展开')

    # ── Step 1: 元占位符 — doing sth / do sth ─────────────────
    if re.search(r'\bdoing\s+sth\b', p, re.IGNORECASE):
        p = re.sub(r'\bdoing\s+sth\b', PH_DOING_STH, p, flags=re.IGNORECASE)
        notes.append('doing sth→元占位符')
    if re.search(r'\bdo\s+sth\b', p, re.IGNORECASE):
        p = re.sub(r'\bdo\s+sth\b', PH_DO_STH, p, flags=re.IGNORECASE)
        notes.append('do sth→元占位符')

    # ── Step 2: sb's / sb / sth / swh / etc.泛化占位符 ──────────
    p = re.sub(r"\bsb's\b", PH_SBPOS, p, flags=re.IGNORECASE)
    p = re.sub(r'\bsb\b',   PH_SB,    p, flags=re.IGNORECASE)
    p = re.sub(r'\bsth\b',  PH_STH,   p, flags=re.IGNORECASE)
    p = re.sub(r'\bswh\b',  PH_SWH,   p, flags=re.IGNORECASE)
    # 泛化占位符：输入表的 _adj_/_noun_/_adv_/_verb_/_timeunit_ 标记
    p = re.sub(r'\b_adj_\b',      PH_ADJ,      p, flags=re.IGNORECASE)
    p = re.sub(r'\b_noun_\b',     PH_NOUN,     p, flags=re.IGNORECASE)
    p = re.sub(r'\b_adv_\b',      PH_ADV,      p, flags=re.IGNORECASE)
    p = re.sub(r'\b_verb_\b',     PH_VERB,     p, flags=re.IGNORECASE)
    p = re.sub(r'\b_timeunit_\b', PH_TIMEUNIT, p, flags=re.IGNORECASE)

    # ── Step 3: 省略号 ────────────────────────────────────────
    if '...' in p:
        p = re.sub(r'([A-Za-z0-9])(\.\.\.)', r'\1 \2', p)
        p = re.sub(r'(\.\.\.)([A-Za-z0-9])', r'\1 \2', p)
        p = p.replace('...', PH_ELLIPSIS)
        notes.append('省略号')

    # ── Step 4: 结尾标点（扩展支持 . 和 ,）────────────
    # spaCy 会把标点单独切出，norm_text 中为 "... word ."
    # 所以允许标点前有可选空格；同时 . 需转义为 \. 避免通配
    end_punct = ''
    p_stripped = p.rstrip()
    if p_stripped.endswith('?'):
        p, end_punct = p_stripped[:-1].rstrip(), r'\s*\?'
    elif p_stripped.endswith('!'):
        p, end_punct = p_stripped[:-1].rstrip(), r'\s*!'
    elif p_stripped.endswith('.'):
        p, end_punct = p_stripped[:-1].rstrip(), r'\s*\.'
        notes.append('句末点号')
    elif p_stripped.endswith(','):
        p, end_punct = p_stripped[:-1].rstrip(), r'\s*,'
        notes.append('句末逗号')

    # ── Step 5: 数字 ──────────────────────────────────────────
    if re.search(r'\d', p):
        p = re.sub(r'\d+(?:[.,]\d+)*[a-z%]*', '\x00NUM\x00', p)
        notes.append('数字')

    # ── Step 5b: 连续数字占位符折叠 ──────────────────────────
    # 处理 "up to 10, 20" / "up to 10 20" 类型：
    # 两个或以上相邻的 NUM（中间可有逗号/空格）折叠为单个 NUM，
    # 使生成的正则匹配任意单个数字（而非两个固定数字）。
    if p.count('\x00NUM\x00') >= 2:
        _ph  = '\x00NUM\x00'
        _pat = (
            r'\x00NUM\x00'
            r'(?:\s*,\s*|\s+)'
            r'\x00NUM\x00'
            r'(?:(?:\s*,\s*|\s+)\x00NUM\x00)*'
        )
        p = re.sub(_pat, lambda _: _ph, p)

    # ── Step 5.5: 被动态检测（在 escape 前，占位符已到位）────────
    # 检测形如 "VERB PH_SB [rest]" 或 "VERB ... PH_SB"（末尾）模式
    # 预处理器会将 assigned→assign、was→be，所以被动分支直接用动词原形
    # 注：不限制 verb 必须在 VERB_FORMS 中，避免遗漏低频动词
    _passive_verb      = None
    _passive_sb_pos    = None   # 'second' | 'last'
    _pt = [t for t in p.split(' ') if t]
    if (len(_pt) >= 2
            and _pt[0]                               # 首词非空
            and _pt[0].isalpha()                     # 首词为纯字母（动词形式）
            and '\x00' not in _pt[0]):               # 首词非占位符
        if _pt[1] == PH_SB:
            _passive_verb   = _pt[0].lower()
            _passive_sb_pos = 'second'
            notes.append('含被动备选')
        elif _pt[-1] == PH_SB and len(_pt) >= 3:
            _passive_verb   = _pt[0].lower()
            _passive_sb_pos = 'last'
            notes.append('含被动备选')

    # ── Step 6: 逐词转义 ──────────────────────────────────────
    tokens = p.split(' ')
    result_tokens = []
    for token in tokens:
        if not token:
            continue
        if '\x00' in token:
            # -ed 备选占位符还原为 (?:ed_form|base_form)
            _alt_m = re.match(r'\x00EDALTS\x00(.+?)\x00ENDALT\x00', token)
            if _alt_m:
                _parts = _alt_m.group(1).split('|')
                result_tokens.append('(?:' + '|'.join(escape_for_regex(pt) for pt in _parts) + ')')
            else:
                result_tokens.append(token)  # 其他占位符直接保留
            continue
        # 枚举标记内容保护：含枚举占位符的 token 跳过 escape
        if ('\x01EL\x01' in token or '\x01ER\x01' in token
                or '\x01EP\x01' in token or '\x01EW\x01' in token):
            result_tokens.append(token)
            continue
        # 连字符词：spaCy 会把 double-edged 拆成 double - edged，
        # 用 \s*-\s* 兼容有无空格两种形式
        if '-' in token and not token.startswith('-') and not token.endswith('-'):
            parts = token.split('-')
            result_tokens.append(r'\s*-\s*'.join(escape_for_regex(pt) for pt in parts))
        else:
            result_tokens.append(escape_for_regex(token))
    p = ' '.join(result_tokens)

    # ── Step 7: 恢复占位符 ────────────────────────────────────
    # 省略号：中间通配最多 6 词，避免跨句连接不相关 token。用 \S+ 而非 \w+，
    # 以匹配归一化文本里作为独立 token 的标点（"hand , i"），\w+ 匹配不到标点。
    p = p.replace(PH_ELLIPSIS,   r'(?:\S+(?:\s+\S+){0,5})?')

    # sb's：覆盖三种归一化形态
    #   "your disposal"             ← 简单物主代词归一（无 's）
    #   "your father 's footsteps"  ← 物主 + 名词 + 's token
    #   "the company 's disposal"   ← 定冠词 + 名词 + 's token
    #   "your interests"            ← your + 复数名词（无 's）
    p = p.replace(PH_SBPOS, _SBPOS_RE)

    # sb 占位符：允许最多2词（如 "his father"）。sb 首词加 _SB_NOT_VERB 约束，
    # 排除高频动词原形，避免 "without having/getting" 等 prep+动名词误配为 sb
    # （动名词经预处理已还原为原形，故排除原形而非 -ing）。
    p = p.replace(PH_SB,         r'(?:_sb_|you|i\b|' + _SB_NOT_VERB + r'\w+(?:\s+\w+)?)')
    # sth 首词排除 place/part：阻止 "take sth in"→"take place in"、
    # "take sth on"→"take place on" 跨"take place"短语误配。
    p = p.replace(PH_STH,        r'(?!(?:place|part)\b)\w+(?:\s+\w+){0,2}')
    p = p.replace(PH_SWH,        r'\w+(?:\s+\w+)?')
    p = p.replace(PH_DO_STH,     _PH_DO_STH_RE)
    # doing sth：预处理器已把 gerund 还原为原形（talking→talk），归一化文本无
    # -ing 后缀，故用 _PH_DOING_STH_RE（负向前瞻 + \w+）匹配原形并排除功能词。
    p = p.replace(PH_DOING_STH,  _PH_DOING_STH_RE)
    p = p.replace('\x00HADBETTER\x00', r'(?:have|would)\s+better')
    p = p.replace('\x00NUM\x00',  _NUM_RE)   # 同时匹配阿拉伯数字和英文数词
    # 恢复 USEDTO / BEGOINGTO 占位符（对应 Pre-step 0.5-pre）
    p = p.replace('\x00USEDTO\x00', r'used?\s+to')
    p = p.replace('\x00BEGOINGTO\x00',
                  r'(?:be|is|are|was|were|am|been)\s+going\s+to')

    # 恢复枚举边界标记为正则语法（对应 Step 0c 高风险结构）
    if '\x01EL\x01' in p:
        def _restore_enum_mark(m):
            words = m.group(1).split('\x01EP\x01')
            parts = []
            for w in words:
                if not w: continue
                if '-' in w:
                    # 含连字符的词（make-up）：直接字面，不加 s?
                    # 注意：不用 re.escape，直接写入正则字符串（连字符在字符类外不需要转义）
                    # 连字符词直接字面写入，只转义真正需要的字符（非字母数字连字符外的特殊符号）
                    safe = re.sub(r'[^a-zA-Z0-9-]', lambda x: re.escape(x.group()), w)
                    parts.append(safe)
                elif re.match(r'^[a-z]+$', w):
                    # 纯字母词：归一化为词干（去掉末尾复数 s）加 s? 匹配单复数
                    # 用 NLTK lemmatizer 思路：末尾是 s 且去掉后≥3字母，则去掉
                    # 避免 notes→not 的问题：只去掉末尾单个 s（不去 es）
                    if w.endswith('s') and len(w) > 3 and not w.endswith('ss'):
                        stem = w[:-1]   # notes→note, records→record
                    else:
                        stem = w        # arm, leg, speech 等不变
                    parts.append(re.escape(stem) + 's?')
                else:
                    parts.append(re.escape(w))
            return '(?:' + '|'.join(parts) + ')'
        p = re.sub(r'\x01EL\x01(.*?)\x01ER\x01', _restore_enum_mark, p)
    p = p.replace('\x01EW\x01', r'\w+')   # 宽槽（etc. 有名词锚定）→ \w+

    p = p.replace(PH_ADJ,      r'\w+')
    p = p.replace(PH_NOUN,     r'\w+')
    p = p.replace(PH_ADV,      r'\w+')
    p = p.replace(PH_VERB,     r'\w+')
    p = p.replace(PH_TIMEUNIT, _TIMEUNIT_RE)

    # ── Step 8: 空格 → \s+ ────────────────────────────────────
    p = re.sub(r' +', r'\\s+', p)

    # ── Step 8b: 标点前 \s+ → \s*（宽松化，兼容 Step 8 清理后的 norm）
    p = re.sub(r'\\s\+([.,;:])', r'\\s*\1', p)

    # ── Step 9: 组装 ──────────────────────────────────────────
    # 末尾标点设为可选（学习者例句常省略标点）。
    _end_suffix = (r'(?:' + end_punct + r')?' if end_punct else '')
    # 开头 be 的短语：谓语前加系动词前缀（对应 Step 0b）。用全形态并列
    # (?:be|is|are|was|...) 而非仅字面 be——预处理器对 "are being"、部分
    # "are + 副词" 语境不稳定归一为 be，全形态可两头兼容。
    _BE_GAP  = 2
    _BE_COPULA = r'(?:be|is|are|was|were|am|been|being)'
    _be_prefix = (r'%s(?:\s+\w+){0,%d}\s+' % (_BE_COPULA, _BE_GAP)) if _leading_be else ''
    regex = r'(?i)\b' + _be_prefix + p + (r'\b' if not end_punct else '') + _end_suffix

    # ── Step 9b: 不定冠词泛化 ─────────────────────────────────
    # a few / a bit / a while 等固定搭配中 a 是短语核心，不应泛化。
    _FIXED_A_PHRASES = {
        'a few', 'a bit', 'a lot', 'a while', 'a minute',
        'a moment', 'a second', 'a little', 'a great deal',
    }
    _orig_lower = phrase.strip().lower()
    _skip_det = any(_orig_lower == fp or _orig_lower.startswith(fp + ' ')
                    for fp in _FIXED_A_PHRASES)
    if not _skip_det:
        regex = re.sub(r'(?<=\\b)an?(?=\\s\+|\\b)',   _DET, regex)
        regex = re.sub(r'(?<=\\s\+)an?(?=\\s\+|\\b)', _DET, regex)

    # ── Step 9b-2: 冠词可选化────────────────────────
    # 将短语首位 \b{DET}\s+ 整体变为可选组 \b(?:{DET}\s+)?
    # 使 "(?i)\b(?:a|an|...)\s+broken\s+homes?\b" 同时匹配
    # "from broken homes"（无冠词、复数，配合 Step 9c 的 s? 生效）
    if not _skip_det:
        _det_space = _DET + r'\s+'          # 在 regex 字符串中的字面内容
        _det_opt   = r'(?:' + _det_space + r')?'
        regex = regex.replace(r'\b' + _det_space,   r'\b' + _det_opt)
        regex = regex.replace(r'\s+' + _det_space,  r'\s+' + _det_opt)

    # ── Step 9c: 末尾名词单复数可选────────────────────
    # 辅音+y → -ies（party→parties）；s/sh/ch/x/z → -es（wish→wishes）；其余 -s。
    if not end_punct:
        def _plural(m):
            w = m.group(1)
            if re.search(r'[bcdfghjklmnpqrstvwxz]y$', w):
                return w[:-1] + r'(?:y|ies)' + m.group(2)
            suf = r'(?:es)?' if re.search(r'(?:ss|sh|ch|s|x|z)$', w) else r's?'
            return w + suf + m.group(2)
        regex = re.sub(r'([a-z]{3,})(\\b)$', _plural, regex)

    # ── Step 9d: 被动态备选分支────────────────────────
    # 在 Step 9c 之后，最终 regex 已定型，追加被动分支。
    # 预处理器归一化后：was→be，assigned→assign（词形还原），
    # he/she→_sb_，所以被动分支用动词原形即可。
    # 关键：将 (?i) 提到最外层 (?i)(?:主动|被动)，
    # 避免 (?:...|(?i)...) 中内嵌 (?i) 触发 re.error。
    if _passive_verb:
        _SB_re = r'(?:_sb_|you|i\b|\w+(?:\s+\w+)?)'   # 与 PH_SB 展开式一致
        if _passive_sb_pos == 'second':
            # 主动: (?i)\bVERB\s+SB\s+rest\b
            # 被动: \bSB\s+be\s+VERB\s+rest\b
            _act_start = r'(?i)\b' + _passive_verb + r'\s+' + _SB_re + r'\s+'
            if regex.startswith(_act_start):
                _rest       = regex[len(_act_start):]           # 结尾部分（已含 \b）
                _act_inner  = r'\b' + _passive_verb + r'\s+' + _SB_re + r'\s+' + _rest
                _pass_inner = r'\b' + _SB_re + r'\s+be\s+' + _passive_verb + r'\s+' + _rest
                regex       = r'(?i)(?:' + _act_inner + r'|' + _pass_inner + r')'
        elif _passive_sb_pos == 'last':
            # 主动: (?i)\bVERB\s+MIDDLE\s+SB\b
            # 被动: \bSB\s+be\s+VERB\s+MIDDLE\b
            _sb_tail = r'\s+' + _SB_re + r'\b'
            if regex.endswith(_sb_tail):
                _act_core   = regex[len(r'(?i)\b'):-len(_sb_tail)]  # 中间部分，如 "cut\s+off"
                _act_inner  = r'\b' + _act_core + _sb_tail
                _pass_inner = r'\b' + _SB_re + r'\s+be\s+' + _act_core + r'\b'
                regex       = r'(?i)(?:' + _act_inner + r'|' + _pass_inner + r')'

    seen, unique_notes = set(), []
    for n in notes:
        if n not in seen:
            seen.add(n)
            unique_notes.append(n)

    return regex, '|'.join(unique_notes)


def generate_regex(phrase_str):
    """短语 → 正则的对外入口，返回 (regex_str, note)。短语表仅含多词短语。"""
    return phrase_to_regex(phrase_str)


# ─────────────────────────────────────────────────────────────
# 6. 主流程：生成 + 去重标注 + 去重输出文件
# ─────────────────────────────────────────────────────────────
def process_file(input_path, output_path, dedup_output_path=None):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    stats = defaultdict(int)

    # ── 第一遍：生成正则 ──────────────────────────────────────
    print("第一遍：生成正则表达式...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        phrase_cell    = row[1]
        regex_cell     = row[7]
        note_cell      = row[8]
        phrase         = phrase_cell.value
        existing_regex = regex_cell.value

        stats['total'] += 1

        if existing_regex and str(existing_regex).strip():
            stats['skipped'] += 1
            continue
        if not phrase or not str(phrase).strip():
            continue

        phrase_str = str(phrase).strip()
        regex, note = generate_regex(phrase_str)

        if '省略号'   in note: stats['ellipsis']    += 1
        if '数字'     in note: stats['number']      += 1
        if '缩写展开' in note: stats['contraction'] += 1
        if 'decade'   in note: stats['merged']      += 1
        if '元占位符' in note: stats['meta_ph']     += 1

        regex_cell.value = regex
        existing_note    = note_cell.value or ''
        note_cell.value  = (existing_note + ' | ' + note
                            if existing_note and note else note or existing_note)
        stats['generated'] += 1

    # ── 第二遍：按正则分组，同时收集各组的 CEFR / 释义 ────────
    # 供去重标注和去重输出文件共用
    print("第二遍：去重标注 + 收集分组信息...")

    # regex_str → {
    #   'first_row': int,              # 首次出现行号（1-based）
    #   'entries':   list of dict,     # 每条原始记录
    # }
    regex_groups = {}   # 保持插入顺序

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        phrase_val = row[1].value
        cefr_val   = row[4].value
        pos_val    = row[5].value
        meaning_val= row[6].value
        regex_val  = row[7].value
        original_val = row[0].value

        if not phrase_val or not regex_val:
            continue
        rx = str(regex_val).strip()
        if not rx:
            continue

        entry = {
            'row':      row_idx,
            'phrase':   str(phrase_val).strip(),
            'original': str(original_val or '').strip(),
            'cefr':     str(cefr_val   or '').strip(),
            'pos':      str(pos_val    or '').strip(),
            'meaning':  str(meaning_val or '').strip(),
        }

        if rx not in regex_groups:
            regex_groups[rx] = {'first_row': row_idx, 'entries': [entry]}
        else:
            regex_groups[rx]['entries'].append(entry)
            # 标注重复行
            dup_note = f'重复正则（同第{regex_groups[rx]["first_row"]}行）'
            existing = row[8].value or ''
            row[8].value = (existing + ' | ' + dup_note if existing else dup_note)

    n_dup = sum(len(g['entries']) - 1 for g in regex_groups.values())
    stats['duplicates'] = n_dup
    wb.save(output_path)

    # ── 第三遍：写去重输出文件 ────────────────────────────────
    if dedup_output_path:
        print("第三遍：写去重正则文件...")
        _write_dedup_file(regex_groups, dedup_output_path)

    # ── 统计输出 ──────────────────────────────────────────────
    n_mono  = sum(1 for g in regex_groups.values() if _is_monosemous(g['entries']))
    n_poly  = len(regex_groups) - n_mono

    print()
    print("=" * 52)
    print("  生成完成")
    print("=" * 52)
    print(f"  总行数:              {stats['total']}")
    print(f"  跳过(已有正则):      {stats['skipped']}")
    print(f"  新生成正则:          {stats['generated']}")
    print(f"  ├─ 省略号:           {stats['ellipsis']}")
    print(f"  ├─ 数字通配:         {stats['number']}")
    print(f"  ├─ 缩写展开:         {stats['contraction']}")
    print(f"  ├─ decade合并:       {stats['merged']}")
    print(f"  └─ do/doing元占位符: {stats['meta_ph']}")
    print()
    print(f"  唯一正则数:          {len(regex_groups)}")
    print(f"  重复标注行数:        {n_dup}")
    print(f"  单义短语:            {n_mono}")
    print(f"  多义短语:            {n_poly}")


def _is_monosemous(entries):
    """
    判断一组共享同一正则的短语是否为单义。
    条件：去重后 CEFR 等级集合大小为 1，且 释义集合大小为 1。
    注意：纯数据重复（同一短语出现多次）也算单义；
          真正的多义是 CEFR 或释义不同，说明同一正则对应不同义项。
    """
    cefr_set    = set(e['cefr']    for e in entries if e['cefr'])
    meaning_set = set(e['meaning'] for e in entries if e['meaning'])
    return len(cefr_set) <= 1 and len(meaning_set) <= 1


def _write_dedup_file(regex_groups, path):
    """写去重后的正则文件，每条唯一正则占一行，标注单义/多义。"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    CEFR_ORDER = ['A1', 'A2', 'B1', 'B2', 'C1', 'C2']
    CEFR_COLORS = {
        'A1': 'E8F5E9', 'A2': 'C8E6C9',
        'B1': 'BBDEFB', 'B2': '90CAF9',
        'C1': 'CE93D8', 'C2': 'BA68C8',
        '':   'F5F5F5',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '去重正则表'

    headers = [
        '正则表达式',
        '代表短语',          # 第一条（最低级别/首条）短语
        '所有短语形式',      # 该正则覆盖的全部不重复短语，分号分隔
        'CEFR等级',          # 去重后所有等级，排序后展示；多个则为范围
        '词性',
        '释义',              # 去重后所有不重复释义，分号分隔
        '合并条数',          # 原始行数（含重复）
        '单义/多义',
        '多义说明',          # 多义时列出各义项的 CEFR + 释义，便于人工核查
        '原始短语（原始形式）',
    ]

    # 表头样式
    hdr_fill = PatternFill("solid", fgColor="2D5986")
    hdr_font = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    col_widths = [60, 28, 55, 12, 12, 45, 8, 10, 55, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    mono_count = poly_count = 0

    for ri, (rx, group) in enumerate(regex_groups.items(), start=2):
        entries = group['entries']
        is_mono = _is_monosemous(entries)

        # ── 代表短语：取 CEFR 最低的那条，若无 CEFR 则取首条 ──
        def cefr_rank(e):
            return CEFR_ORDER.index(e['cefr']) if e['cefr'] in CEFR_ORDER else 99
        rep_entry = min(entries, key=cefr_rank)

        # ── 所有不重复短语形式（保持首次出现顺序）────────────
        seen_phrases = set()
        uniq_phrases = []
        for e in entries:
            if e['phrase'] not in seen_phrases:
                seen_phrases.add(e['phrase'])
                uniq_phrases.append(e['phrase'])

        # ── CEFR：排序后展示 ──────────────────────────────────
        cefr_set = sorted(
            set(e['cefr'] for e in entries if e['cefr']),
            key=lambda x: CEFR_ORDER.index(x) if x in CEFR_ORDER else 99
        )
        if len(cefr_set) == 0:
            cefr_disp = ''
        elif len(cefr_set) == 1:
            cefr_disp = cefr_set[0]
        else:
            cefr_disp = f"{cefr_set[0]}~{cefr_set[-1]}"

        # ── 词性：去重 ────────────────────────────────────────
        seen_pos = set()
        uniq_pos = []
        for e in entries:
            if e['pos'] and e['pos'] not in seen_pos:
                seen_pos.add(e['pos'])
                uniq_pos.append(e['pos'])

        # ── 释义：去重 ────────────────────────────────────────
        seen_meanings = set()
        uniq_meanings = []
        for e in entries:
            if e['meaning'] and e['meaning'] not in seen_meanings:
                seen_meanings.add(e['meaning'])
                uniq_meanings.append(e['meaning'])

        # ── 原始短语：去重 ────────────────────────────────────
        seen_orig = set()
        uniq_orig = []
        for e in entries:
            if e['original'] and e['original'] not in seen_orig:
                seen_orig.add(e['original'])
                uniq_orig.append(e['original'])

        # ── 多义说明：列出各不重复义项的 CEFR + 释义 ─────────
        if is_mono:
            poly_note = ''
            mono_count += 1
        else:
            # 按 (cefr, meaning) 去重，生成可读说明
            seen_senses = set()
            sense_lines = []
            for e in entries:
                key = (e['cefr'], e['meaning'])
                if key not in seen_senses and (e['cefr'] or e['meaning']):
                    seen_senses.add(key)
                    sense_lines.append(f"[{e['cefr']}] {e['meaning']}")
            poly_note = ' | '.join(sense_lines)
            poly_count += 1

        # ── 行填充颜色：按 CEFR 最低级别 ─────────────────────
        base_cefr = cefr_set[0] if cefr_set else ''
        fill_color = CEFR_COLORS.get(base_cefr, 'F5F5F5')
        # 多义短语加橙色标记覆盖
        if not is_mono:
            fill_color = 'FFE0B2'
        fill = PatternFill("solid", fgColor=fill_color)

        row_data = [
            rx,
            rep_entry['phrase'],
            '; '.join(uniq_phrases),
            cefr_disp,
            '; '.join(uniq_pos),
            '; '.join(uniq_meanings),
            len(entries),
            '单义' if is_mono else '多义',
            poly_note,
            '; '.join(uniq_orig),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(wrap_text=False, vertical='top')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(path)
    print(f"  → {path}（单义 {mono_count} 条，多义 {poly_count} 条）")



# ─────────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import argparse

    ap = argparse.ArgumentParser(
        description='为短语表批量生成正则表达式列，并输出去重正则文件。')
    ap.add_argument('--input',  required=True,
                    help='输入短语表 xlsx（含"拆分后"短语列，正则列为空则生成）')
    ap.add_argument('--output', required=True,
                    help='输出 xlsx（在输入基础上写入"正则表达式"列 + 重复标注）')
    ap.add_argument('--dedup-output', default=None,
                    help='可选：去重正则文件 xlsx（每条唯一正则一行，标注单义/多义）')
    args = ap.parse_args()

    process_file(args.input, args.output, dedup_output_path=args.dedup_output)
