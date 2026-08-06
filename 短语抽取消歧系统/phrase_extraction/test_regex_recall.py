#!/usr/bin/env python3
"""
正则召回率测试（例句集口径）
============================
沿用历史 retester 方法：对每个短语，把它的【词典例句】【学习者例句】用生产
预处理器(FixedPreprocessor)归一化，再用该短语的【生产正则】search，命中即召回。
按原始短语词条分组（任意变体命中即算该词条召回），分来源(词典/学习者)与 CEFR 统计。

召回 ≠ 精确率：本测只测"regex 能否在短语自己的例句里命中自己"（漏不漏），
不测假阳（那要用作文语料 + 人工标注，见评测集 P/R）。

用法：
  python test_regex_recall.py [--table data/单义短语_v12_regex.xlsx]
"""
import argparse, re
from collections import defaultdict
import openpyxl
from extractor import FixedPreprocessor
from evp_regex_generator import generate_regex

_pre = FixedPreprocessor()


def test_sents(compiled, raw_text):
    """raw_text 内每句(按换行拆)归一化后 search。返回 any_hit|None（无例句 None）。"""
    if not raw_text:
        return None
    any_hit = False
    for s in re.split(r'\n+', raw_text):
        s = s.strip()
        if not s:
            continue
        try:
            norm = _pre.normalize(s)
        except Exception:
            norm = s.lower()
        if compiled.search(norm):
            any_hit = True
    return any_hit


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--table', default='data/单义短语_v12_regex.xlsx')
    ap.add_argument('--stored', action='store_true',
                    help='用表内存档正则（默认用当前生成器重新生成，反映最新改动）')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.table, read_only=True); ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    C = {v: i for i, v in enumerate(h)}
    need = ['原始短语', '拆分后', 'CEFR等级', '正则表达式', '词典例句', '学习者例句']
    miss = [c for c in need if c not in C]
    if miss:
        raise SystemExit(f'缺列：{miss}')

    # 逐变体：用正则测两类例句（默认重新生成，反映生成器最新改动）
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        phrase = str(row[C['拆分后']] or '').strip()
        if not phrase or len(phrase.split()) < 2:   # 单词条目不计（非 MWE）
            continue
        if args.stored:
            regex = str(row[C['正则表达式']] or '').strip()
        else:
            try:
                regex, _ = generate_regex(phrase)
            except Exception:
                regex = ''
        if not regex or regex in ('None', 'nan'):
            continue
        try:
            compiled = re.compile(regex)
        except re.error:
            compiled = None
        dict_ex  = str(row[C['词典例句']] or '').strip()
        learn_ex = str(row[C['学习者例句']] or '').strip()
        records.append({
            'orig': str(row[C['原始短语']] or '').strip(),
            'cefr': str(row[C['CEFR等级']] or '').strip(),
            'regex': regex, 'err': compiled is None,
            'dict_ex': dict_ex, 'learn_ex': learn_ex,
            'dict_hit':  test_sents(compiled, dict_ex)  if compiled is not None else None,
            'learn_hit': test_sents(compiled, learn_ex) if compiled is not None else None,
        })

    # 按原始短语分组：任意变体命中即算召回；斜杠词条做 H 类跨变体兜底
    groups = defaultdict(list)
    for r in records:
        groups[r['orig']].append(r)

    d_tot = d_hit = l_tot = l_hit = 0
    n_pass = n_fail = n_err = n_cross = 0
    cefr = defaultdict(lambda: {'d_t': 0, 'd_h': 0, 'l_t': 0, 'l_h': 0})
    fails = []

    for orig, recs in groups.items():
        has_valid = any(not r['err'] for r in recs)
        g_dict_hit  = any(r['dict_hit']  for r in recs if r['dict_hit']  is not None)
        g_learn_hit = any(r['learn_hit'] for r in recs if r['learn_hit'] is not None)
        any_hit = g_dict_hit or g_learn_hit

        # H 类：斜杠词条各变体共享例句，合并归一化文本后任一变体正则命中即算
        cross = False
        if '/' in orig and has_valid and not any_hit:
            combined = ' '.join(
                (_safe_norm(r['dict_ex']) + ' ' + _safe_norm(r['learn_ex']))
                for r in recs if not r['err']).lower()
            for r in recs:
                if not r['err'] and re.search(r['regex'], combined):
                    cross = True; break

        if not has_valid:      n_err  += 1
        elif any_hit or cross:
            n_pass += 1
            if cross: n_cross += 1
        else:
            n_fail += 1
            fails.append((orig, recs[0]['cefr']))

        lv = recs[0]['cefr']
        if any(r['dict_ex'] for r in recs):
            d_tot += 1; cefr[lv]['d_t'] += 1
            if g_dict_hit or cross: d_hit += 1; cefr[lv]['d_h'] += 1
        if any(r['learn_ex'] for r in recs):
            l_tot += 1; cefr[lv]['l_t'] += 1
            if g_learn_hit or cross: l_hit += 1; cefr[lv]['l_h'] += 1

    pc = lambda a, b: f'{a/b*100:.1f}%' if b else '—'
    print(f'\n词条组数 {len(groups)} | PASS {n_pass}  FAIL {n_fail}  REGEX_ERROR {n_err}'
          f'  (其中 H类跨变体兜底 {n_cross})')
    print(f'总体词条召回(有任一类例句且命中): {pc(n_pass, n_pass + n_fail)}')
    print(f'\n分来源召回（按原始短语词条）:')
    print(f'  词典例句 : {d_hit}/{d_tot} = {pc(d_hit, d_tot)}')
    print(f'  学习者例句: {l_hit}/{l_tot} = {pc(l_hit, l_tot)}')
    print(f'\n分 CEFR（词典 / 学习者）:')
    for lv in ['A1', 'A2', 'B1', 'B2', 'C1', 'C2', '']:
        g = cefr.get(lv)
        if not g or (g['d_t'] == 0 and g['l_t'] == 0):
            continue
        print(f'  {lv or "(无级)":6} 词典 {g["d_h"]:4}/{g["d_t"]:<4} {pc(g["d_h"],g["d_t"]):>7}'
              f'  | 学习者 {g["l_h"]:4}/{g["l_t"]:<4} {pc(g["l_h"],g["l_t"]):>7}')
    print(f'\n未命中词条 {len(fails)} 例（前 30）:')
    for orig, lv in fails[:30]:
        print(f'  [{lv:3}] {orig}')


def _safe_norm(t):
    if not t:
        return ''
    try:
        return ' '.join(_pre.normalize(s) for s in re.split(r'\n+', t) if s.strip())
    except Exception:
        return t.lower()


if __name__ == '__main__':
    main()
