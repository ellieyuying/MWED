#!/usr/bin/env python3
"""
EVP 全流水线：抽取 + 消歧（用最新 v12 词表）
============================================
串联两个独立模块：
  phrase_extraction/extractor.py   —— 抽取（作文→句子→匹配）
  disambiguation/disambiguator.py  —— 多义消歧（LLM 判义项）

输出：
  · sentences.xlsx   句子表（每句一行）
  · hit_detail.xlsx  抽取+消歧结果（每命中一行，含义项/置信度/理由）

用法：
  # 消歧需 API key（环境变量），dry-run 免 key。模型默认 qwen3.6-flash。
  export DASHSCOPE_API_KEY=...
  python run_pipeline.py --essays-dir <XML目录> --out-dir 结果_v12
  # 复用已有义项标注（仅新命中调 LLM）：
  python run_pipeline.py --essays-dir <XML目录> --out-dir 结果_v12 --reuse-disambig 旧hit_detail.xlsx
  # 免 API 试跑：--dry-run
"""
import sys, os, argparse
from pathlib import Path

_ROOT = Path(__file__).parent
sys.path.insert(0, str(_ROOT / 'phrase_extraction'))
sys.path.insert(0, str(_ROOT / 'disambiguation'))

import extractor as EX
import disambiguator as DIS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _load_disambig_cache(path):
    """旧 hit_detail → {(短语, 句子): {'sense','confidence','reason'}}，供复用。"""
    wb = openpyxl.load_workbook(path, read_only=True); ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    I = {x: i for i, x in enumerate(h)}
    cache = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if str(r[I['多义']]).strip() != '是':
            continue
        key = (str(r[I['短语']]).strip(), str(r[I['句子']]).strip())
        cache[key] = {'sense': r[I['消歧义项']], 'confidence': r[I['置信度']],
                      'reason': r[I['消歧理由']]}
    wb.close()
    return cache


def run(essay_paths, unambig, ambig, out_dir, dry_run=False, reuse_path=None):
    os.makedirs(out_dir, exist_ok=True)

    # ── 1. 抽取 ────────────────────────────────────────────────
    print(f'[1/3] 抽取：{len(essay_paths)} 个文件 ...')
    results = EX.extract(essay_paths, unambig, ambig)   # [(Answer, [MatchHit])]
    n_ans  = len(results)
    n_hits = sum(len(h) for _, h in results)
    print(f'      answer 数：{n_ans} | 命中总数：{n_hits}')

    # ── 2. 跨篇批量消歧（只对多义命中）─────────────────────────
    flat = [(a, h) for a, hits in results for h in hits]
    ambig_pos = [k for k, (_, h) in enumerate(flat) if h.pattern.is_ambiguous]
    instances = [{
        'phrase':  flat[k][1].pattern.phrase_id,
        'target':  flat[k][1].sentence,
        'context': flat[k][1].context,
        'senses':  flat[k][1].pattern.senses,
    } for k in ambig_pos]

    # 复用旧义项标注：(短语, 句子) 命中即复用，仅未命中的调 LLM
    cache = _load_disambig_cache(reuse_path) if reuse_path else {}
    reuse_by_i, todo_idx, todo_inst = {}, [], []
    for i, inst in enumerate(instances):
        hit = cache.get((str(inst['phrase']).strip(), str(inst['target']).strip()))
        if hit is not None:
            reuse_by_i[i] = hit
        else:
            todo_idx.append(i); todo_inst.append(inst)
    n_reuse = len(reuse_by_i)
    print(f'[2/3] 消歧：多义命中 {len(ambig_pos)} 条'
          + (f'，复用旧标注 {n_reuse}，需 LLM {len(todo_inst)}' if reuse_path else '')
          + (' (dry-run 占位)' if dry_run
             else f'，约 {(len(todo_inst)+DIS.BATCH_SIZE-1)//DIS.BATCH_SIZE} 次 API'))
    fresh = DIS.disambiguate(todo_inst, dry_run=dry_run) if todo_inst else []
    for j, i in enumerate(todo_idx):
        reuse_by_i[i] = fresh[j]
    sense_out = [reuse_by_i[i] for i in range(len(instances))]
    result_by_pos = {k: o for k, o in zip(ambig_pos, sense_out)}

    # ── 3. 写输出 ──────────────────────────────────────────────
    print('[3/3] 写输出 ...')
    EX.write_sentences(results, str(Path(out_dir) / 'sentences.xlsx'))
    _write_hit_detail(flat, result_by_pos, str(Path(out_dir) / 'hit_detail.xlsx'))


def _write_hit_detail(flat, result_by_pos, out_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '抽取+消歧'
    hdr = ['essay_id', '题号', '总分', '句序', '句子', '短语', '变体形式',
           'CEFR', '词性', '释义', '多义', '消歧义项', '置信度', '消歧理由']
    ws.append(hdr)
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='DDEEFF')
    for k, (a, h) in enumerate(flat):
        o = result_by_pos.get(k)
        if h.pattern.is_ambiguous:
            sense = str(o.get('sense', '')) if o else ''
            conf  = str(o.get('confidence', '')) if o else ''
            reason= str(o.get('reason', '')) if o else ''
        else:
            sense, conf, reason = '1', '', ''   # 单义：义项恒为1
        ws.append([a.essay_id, a.question_no, a.total_score, h.sent_idx, h.sentence,
                   h.pattern.phrase_id, h.pattern.phrase_form, h.pattern.cefr,
                   h.pattern.pos, h.pattern.meaning,
                   '是' if h.pattern.is_ambiguous else '否', sense, conf, reason])
    for i, w in enumerate([22,6,6,6,55,24,22,7,7,36,6,8,8,40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row: c.alignment = Alignment(vertical='top', wrap_text=True)
    wb.save(out_path)
    print(f'[保存] hit_detail → {out_path}')


def main():
    ap = argparse.ArgumentParser(description='EVP 抽取+消歧全流水线（v12）')
    ap.add_argument('--essays', nargs='+')
    ap.add_argument('--essays-dir')
    ap.add_argument('--unambig', default=str(_ROOT / 'phrase_extraction' / 'data' / '单义短语_v12_regex.xlsx'))
    ap.add_argument('--ambig',   default=str(_ROOT / 'phrase_extraction' / 'data' / '多义短语_v10.xlsx'))
    ap.add_argument('--out-dir', default='结果_v12')
    ap.add_argument('--dry-run', action='store_true', help='消歧不调 API（占位）')
    ap.add_argument('--reuse-disambig', default=None,
                    help='旧 hit_detail.xlsx：(短语,句子) 命中即复用义项，仅未命中调 LLM')
    args = ap.parse_args()

    paths = list(args.essays or [])
    if args.essays_dir:
        paths += [str(p) for p in Path(args.essays_dir).rglob('*.xml')]
    if not paths:
        ap.error('请指定 --essays 或 --essays-dir')
    run(paths, args.unambig, args.ambig, args.out_dir,
        dry_run=args.dry_run, reuse_path=args.reuse_disambig)


if __name__ == '__main__':
    main()
