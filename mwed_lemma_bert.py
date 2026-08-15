#!/usr/bin/env python3
"""
EVP 多义短语消歧 —— 嵌入相似度 Baseline
========================================
支持四种编码器（通过命令行参数切换）：
  --encoder tfidf          : TF-IDF + cosine（无需联网，可直接运行）
  --encoder bert           : sentence-transformers/all-MiniLM-L6-v2（原版小模型）
  --encoder bert_sentence  : bert-base-uncased，整句 mean-pool 向量
  --encoder bert_phrase    : bert-base-uncased，仅取目标短语 token 的上下文向量

用法示例
--------
  python wsd_baseline_minilm.py --encoder bert_sentence
  python D:\projects\MWED\dab\wsd_lemma_bert.py --encoder bert_phrase


注：bert_phrase 模式下 --strategy 参数无效；
    义项锚点始终从词典例句中提取短语向量，无例句时退化为整句向量。

输入
----
  多义短语.xlsx  （与本脚本同目录，或通过 --data 指定路径）

输出
----
  wsd_eval_results.xlsx   各实例预测结果 + 汇总
"""

import argparse, re, sys
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sklearn.metrics.pairwise import cosine_similarity


# ══════════════════════════════════════════════════════════════
# 1. 数据加载
# ══════════════════════════════════════════════════════════════

def clean_text(text: str) -> str:
    """去掉 [correction] 括号注释，合并多余空白"""
    if not text:
        return ""
    text = re.sub(r'\[.*?\]', '', str(text))
    return re.sub(r'\s+', ' ', text).strip()


def load_phrase_data(xlsx_path: str) -> dict:
    """
    读取多义短语 Excel，返回：
      {词条: [sense_dict, ...]}
    其中 sense_dict 保留 Excel 原始字段，额外加 clean_learner。
    去重规则：以 (词条, 释义序号) 为主键，忽略因中心词不同导致的重复行。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    seen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        key = (r['词条'], r['释义序号'])
        if key not in seen:
            r['clean_learner'] = clean_text(r.get('学习者例句', ''))
            seen[key] = r

    phrase_data: dict = defaultdict(list)
    for r in seen.values():
        phrase_data[r['词条']].append(r)

    # 按释义序号排序，保证义项顺序稳定
    for p in phrase_data:
        phrase_data[p].sort(key=lambda x: x['释义序号'])

    return dict(phrase_data)


def filter_evaluable(phrase_data: dict) -> dict:
    """
    保留满足评估条件的短语：
      - 义项数 >= 2
      - 每个义项都有学习者例句
    """
    return {
        p: senses for p, senses in phrase_data.items()
        if len(senses) >= 2 and all(s['clean_learner'] for s in senses)
    }


# ══════════════════════════════════════════════════════════════
# 2. 义项文本构建（仅用于 tfidf / bert / bert_sentence 模式）
# ══════════════════════════════════════════════════════════════

def build_sense_text(sense: dict, strategy: str) -> str:
    """
    strategy:
      'def'    : 仅释义（~13词）
      'def+ex' : 释义 + 词典例句（更丰富的语义锚点）
    """
    parts = [clean_text(sense.get('释义', ''))]
    if strategy == 'def+ex' and sense.get('词典例句'):
        parts.append(clean_text(sense['词典例句']))
    return ' '.join(p for p in parts if p)


# ══════════════════════════════════════════════════════════════
# 3. 编码器（可插拔）
# ══════════════════════════════════════════════════════════════

class TFIDFEncoder:
    """TF-IDF + cosine，完全离线"""

    def __init__(self):
        from sklearn.feature_extraction.text import TfidfVectorizer
        self._cls = TfidfVectorizer

    def encode(self, texts: list[str]) -> np.ndarray:
        vec = self._cls(stop_words=None)
        return vec.fit_transform(texts).toarray()

    @property
    def name(self):
        return "TF-IDF"


class BERTEncoder:
    """原版：sentence-transformers 小模型（all-MiniLM-L6-v2，约 80MB）"""

    def __init__(self, model_name: str = 'all-MiniLM-L6-v2'):
        try:
            from sentence_transformers import SentenceTransformer
        except ImportError:
            sys.exit("请先安装：pip install sentence-transformers")
        print(f"[INFO] 加载 BERT 模型: {model_name}  (首次运行需下载，约 80MB)")
        self._model = SentenceTransformer(model_name)
        self._model_name = model_name

    def encode(self, texts: list[str]) -> np.ndarray:
        return self._model.encode(texts, show_progress_bar=False)

    @property
    def name(self):
        return f"BERT ({self._model_name})"


# ── 新增 ① bert_sentence ────────────────────────────────────────
class BERTSentenceEncoder:
    """
    bert-base-uncased + mean pooling，整句向量。
    与 all-MiniLM-L6-v2 的区别：
      - 原生 BERT，未针对句子相似度任务微调
      - 隐藏维度 768（vs MiniLM 的 384）
      - 约 400MB，精度更高但速度较慢
    用法：--encoder bert_sentence
    """

    def __init__(self, model_name: str = 'bert-base-uncased'):
        try:
            from transformers import AutoTokenizer, AutoModel
            import torch
        except ImportError:
            sys.exit("请先安装：pip install transformers torch")
        print(f"[INFO] 加载 BERT-sentence 模型: {model_name}  (首次运行需下载，约 400MB)")
        self._tok   = AutoTokenizer.from_pretrained(model_name)
        self._model = AutoModel.from_pretrained(model_name)
        self._model.eval()
        self._torch = torch
        self._model_name = model_name

    def _mean_pool(self, last_hidden_state, attention_mask) -> np.ndarray:
        """对非 padding token 做加权平均"""
        mask = attention_mask.unsqueeze(-1).float()          # (1, seq, 1)
        summed   = (last_hidden_state * mask).sum(dim=1)     # (1, hidden)
        counts   = mask.sum(dim=1).clamp(min=1e-9)           # (1, 1)
        return (summed / counts).squeeze(0).detach().numpy() # (hidden,)

    def encode(self, texts: list[str]) -> np.ndarray:
        vecs = []
        for t in texts:
            inputs = self._tok(
                t, return_tensors='pt',
                truncation=True, max_length=512, padding=True
            )
            with self._torch.no_grad():
                out = self._model(**inputs)
            vecs.append(self._mean_pool(out.last_hidden_state, inputs['attention_mask']))
        return np.array(vecs)

    @property
    def name(self):
        return f"BERT-sentence ({self._model_name})"


# ── 新增 ② bert_phrase ─────────────────────────────────────────
# 短语中可能出现的占位符，不作为匹配关键词
_PH_SKIP = frozenset({
    'sb', 'sth', 'swh',
    'do', 'doing', 'be', 'have', 'not',
    'to', 'a', 'an', 'the',
    'your', 'yourself', 'one',
})


def _phrase_keywords(phrase: str) -> list[str]:
    """
    从短语字符串中提取用于定位的关键词（去掉占位符和虚词）。
    示例：
      "make up for sth"   → ['make', 'up', 'for']
      "be in sb's shoes"  → ['shoes']
      "go on"             → ['go', 'on']
    """
    return [w for w in re.sub(r"['\-]", ' ', phrase.lower()).split()
            if w not in _PH_SKIP]


class BERTPhraseEncoder:
    """
    bert-base-uncased + 短语 token 上下文向量。

    核心思路（同 Hu et al. 2025 / Lu & Hu 2022）：
      - 把整句送入 BERT，取 last_hidden_state
      - 只提取目标短语对应 token 的向量并做平均
      - 得到的向量携带整句上下文信息，但聚焦于短语本身

    义项锚点构建（在 disambiguate 中处理）：
      - 对每条词典例句提取短语向量，多句取平均 → 义项向量
      - 无例句时退化为整句 mean-pool

    用法：--encoder bert_phrase
    注：--strategy 参数在此模式下无效。
    """

    def __init__(self, model_name: str = 'bert-base-uncased',
                 spacy_model: str = 'en_core_web_sm'):
        try:
            from transformers import AutoTokenizer, AutoModel
            import torch
        except ImportError:
            sys.exit("请先安装：pip install transformers torch")
        print(f"[INFO] 加载 BERT-phrase 模型: {model_name}  (首次运行需下载，约 400MB)")
        self._tok   = AutoTokenizer.from_pretrained(model_name)
        self._model = AutoModel.from_pretrained(model_name)
        self._model.eval()
        self._torch = torch
        self._model_name = model_name
        self._keywords: list[str] = []   # 当前短语关键词，由 set_phrase() 设置
        self._match_log: list[dict] = [] # 消融记录：每次编码的匹配情况

        # ── spaCy 词形还原（可选，失败时退化为精确匹配）──────────────
        try:
            import spacy as _sp
            self._nlp = _sp.load(spacy_model)
            self._spacy_model = spacy_model
            print(f"[INFO] spaCy 词形还原已启用: {spacy_model}")
        except Exception as _e:
            self._nlp = None
            self._spacy_model = None
            print(f"[WARN] spaCy 未加载（{_e}），退化为原始词形匹配")

    def set_phrase(self, phrase: str):
        """在处理每条短语前调用，更新关键词列表"""
        self._keywords = _phrase_keywords(phrase)
        if not self._keywords:
            print(f"[WARN] 短语 '{phrase}' 提取不到关键词，将退化为整句向量")

    # ── 核心：在 offset_mapping 中定位短语 token ──────────────────
    # ── 字符区间 → BERT token 下标（内部工具）──────────────────────
    def _span_to_bert_indices(self, char_start: int, char_end: int,
                               offset_mapping: list) -> list[int]:
        """把字符区间 [char_start, char_end) 映射到 BERT token 下标列表。"""
        indices = []
        for idx, (tok_s, tok_e) in enumerate(offset_mapping):
            tok_s, tok_e = int(tok_s), int(tok_e)
            if tok_s == 0 and tok_e == 0:
                continue   # [CLS]/[SEP]
            if tok_s < char_end and tok_e > char_start:
                indices.append(idx)
        return indices

    def _find_exact(self, text: str, offset_mapping: list) -> list[int]:
        """原始精确字符匹配（不做词形还原）。"""
        text_lower = text.lower()
        kw_spans: list[tuple[int, int]] = []
        search_from = 0
        for kw in self._keywords:
            pos = text_lower.find(kw, search_from)
            if pos == -1:
                pos = text_lower.find(kw)
            if pos == -1:
                return []
            kw_spans.append((pos, pos + len(kw)))
            search_from = pos + 1
        char_start = min(s for s, _ in kw_spans)
        char_end   = max(e for _, e in kw_spans)
        return self._span_to_bert_indices(char_start, char_end, offset_mapping)

    def _find_by_lemma(self, text: str, offset_mapping: list) -> list[int]:
        """
        spaCy 词形还原匹配：
          1. 用 spaCy 把句子每个 token 还原成原形
          2. 用还原后的词形匹配短语关键词（关键词本身已是原形）
          3. 取匹配 token 在原始文本中的字符位置
          4. 映射到 BERT token 下标
        例：短语 'make up'，句子 'She made up her mind'
            spaCy lemma: made→make ✓，直接命中
        """
        doc = self._nlp(text)
        # (lemma, 原始小写, 字符起始, 字符结束)
        sp_toks = [(t.lemma_.lower(), t.text.lower(), t.idx, t.idx + len(t.text))
                   for t in doc]

        kw_char_spans: list[tuple[int, int]] = []
        search_from = 0
        for kw in self._keywords:
            found = False
            # 先按顺序找（保证短语词序）
            for i in range(search_from, len(sp_toks)):
                lemma, orig, cs, ce = sp_toks[i]
                if lemma == kw or orig == kw:
                    kw_char_spans.append((cs, ce))
                    search_from = i + 1
                    found = True
                    break
            if not found:
                # 放宽：不限顺序（应对被动语态等词序变化）
                for lemma, orig, cs, ce in sp_toks:
                    if lemma == kw or orig == kw:
                        kw_char_spans.append((cs, ce))
                        found = True
                        break
            if not found:
                return []

        char_start = min(s for s, _ in kw_char_spans)
        char_end   = max(e for _, e in kw_char_spans)
        return self._span_to_bert_indices(char_start, char_end, offset_mapping)

    def _find_token_indices(self, text: str, offset_mapping: list) -> list[int]:
        """
        主入口：优先用 spaCy lemma 匹配，不可用时退化为精确匹配。
        返回覆盖目标短语关键词的 BERT token 下标列表。
        """
        if not self._keywords:
            return []
        if self._nlp is not None:
            indices = self._find_by_lemma(text, offset_mapping)
            if indices:
                return indices
            # lemma 匹配失败（极少数情况）→ 精确匹配兜底
        return self._find_exact(text, offset_mapping)

    # ── 单句编码：返回短语向量（找不到时退化为整句 mean-pool）────────
    def encode_one(self, text: str, label: str = '') -> np.ndarray:
        """
        编码单句并记录匹配情况到 _match_log。
        label: 调用方标记，如 'target'（学习者例句）或 'sense_ex'（义项例句），
               便于后续分组分析。
        """
        inputs = self._tok(
            text, return_tensors='pt',
            truncation=True, max_length=512,
            return_offsets_mapping=True
        )
        offset_mapping = inputs.pop('offset_mapping')[0].tolist()

        with self._torch.no_grad():
            out = self._model(**inputs)
        hidden = out.last_hidden_state[0]   # (seq_len, 768)

        indices = self._find_token_indices(text, offset_mapping)
        matched = len(indices) > 0

        # ── 消融记录（新增 match_method 字段）────────────────────
        if matched:
            method = 'lemma' if self._nlp is not None else 'exact'
        else:
            method = 'fallback'
        self._match_log.append({
            'phrase':        ' '.join(self._keywords),
            'label':         label,
            'text':          text[:80],
            'status':        'matched' if matched else 'fallback',
            'match_method':  method,
            'n_tokens':      len(indices),
        })

        if matched:
            return hidden[indices].mean(dim=0).detach().numpy()
        else:
            return hidden[1:-1].mean(dim=0).detach().numpy()

    def encode(self, texts: list[str], label: str = '') -> np.ndarray:
        return np.array([self.encode_one(t, label=label) for t in texts])

    # ── 消融统计辅助方法 ──────────────────────────────────────────
    def reset_log(self):
        self._match_log = []

    def get_match_stats(self) -> dict:
        """
        汇总 _match_log，返回：
          total / matched / fallback / match_rate
          按 label 分组的明细（target vs sense_ex）
          按 phrase 分组的明细（哪些短语匹配率低）
        """
        log = self._match_log
        if not log:
            return {}

        def _stats(subset):
            n = len(subset)
            m = sum(1 for r in subset if r['status'] == 'matched')
            return {'total': n, 'matched': m, 'fallback': n - m,
                    'match_rate': round(m / n, 4) if n else 0}

        by_label: dict = {}
        for r in log:
            by_label.setdefault(r['label'], []).append(r)

        by_phrase: dict = {}
        for r in log:
            by_phrase.setdefault(r['phrase'], []).append(r)

        by_method: dict = {}
        for r in log:
            by_method.setdefault(r.get('match_method', 'unknown'), []).append(r)

        return {
            'overall':   _stats(log),
            'by_label':  {lbl: _stats(rows) for lbl, rows in by_label.items()},
            'by_phrase': {ph:  _stats(rows) for ph,  rows in by_phrase.items()},
            'by_method': {m:   _stats(rows) for m,   rows in by_method.items()},
            'raw_log':   log,
        }

    @property
    def name(self):
        return f"BERT-phrase ({self._model_name})"


# ══════════════════════════════════════════════════════════════
# 4. 消歧核心逻辑
# ══════════════════════════════════════════════════════════════

def disambiguate(target_sentence: str, senses: list[dict], phrase: str,
                 encoder, strategy: str) -> tuple[int, list[float], dict]:
    """
    返回 (预测的释义序号, 各义项相似度列表, 匹配状态字典)

    匹配状态字典（仅 bert_phrase 模式有意义）：
      target_matched  : 学习者例句是否找到了短语 token
      sense_match_rate: 义项例句的平均匹配率（多条例句时）
    """
    match_info = {'target_matched': None, 'sense_match_rate': None}

    if isinstance(encoder, BERTPhraseEncoder):
        encoder.set_phrase(phrase)

        # ── 义项锚点：从词典例句提取短语向量 ──────────────────────
        sense_vecs = []
        sense_match_counts = []
        for s in senses:
            raw_ex = str(s.get('词典例句', '') or '')
            ex_sentences = [clean_text(e)
                            for e in re.split(r'[\n;；]+', raw_ex)
                            if clean_text(e)]
            log_before = len(encoder._match_log)
            if ex_sentences:
                vecs = encoder.encode(ex_sentences, label='sense_ex')
                sense_vecs.append(vecs.mean(axis=0))
            else:
                vec = encoder.encode_one(clean_text(s.get('释义', '')),
                                         label='sense_ex')
                sense_vecs.append(vec)

            # 统计这批例句的匹配率
            new_logs = encoder._match_log[log_before:]
            n_matched = sum(1 for r in new_logs if r['status'] == 'matched')
            sense_match_counts.append(
                n_matched / len(new_logs) if new_logs else 0
            )

        match_info['sense_match_rate'] = round(
            sum(sense_match_counts) / len(sense_match_counts), 4
        ) if sense_match_counts else 0

        # ── 目标向量：从学习者例句提取短语向量 ─────────────────────
        log_before = len(encoder._match_log)
        target_emb = encoder.encode([clean_text(target_sentence)], label='target')
        target_log = encoder._match_log[log_before:]
        match_info['target_matched'] = (
            target_log[0]['status'] == 'matched' if target_log else False
        )

        sense_embs = np.array(sense_vecs)

    else:
        sense_texts = [build_sense_text(s, strategy) for s in senses]
        all_texts   = sense_texts + [clean_text(target_sentence)]
        embeddings  = encoder.encode(all_texts)
        target_emb  = embeddings[-1:]
        sense_embs  = embeddings[:-1]

    sims     = cosine_similarity(target_emb, sense_embs)[0].tolist()
    best_idx = int(np.argmax(sims))
    return senses[best_idx]['释义序号'], sims, match_info


# ══════════════════════════════════════════════════════════════
# 5. 评估
# ══════════════════════════════════════════════════════════════

def evaluate(eval_phrases: dict, encoder, strategy: str) -> tuple[list[dict], dict]:
    """
    对所有可评估实例做消歧预测，返回：
      results : 每条实例的详细记录
      summary : 汇总指标
    """
    results = []
    correct = 0
    total   = 0
    per_n   = defaultdict(lambda: {'correct': 0, 'total': 0})

    for phrase, senses in eval_phrases.items():
        n = len(senses)
        for gold_sense in senses:
            target  = gold_sense['clean_learner']
            gold_no = gold_sense['释义序号']

            pred_no, sims, match_info = disambiguate(
                target, senses, phrase, encoder, strategy
            )
            is_correct = (pred_no == gold_no)

            sorted_sims = sorted(sims, reverse=True)
            sim_gap = sorted_sims[0] - sorted_sims[1] if len(sorted_sims) > 1 else 1.0
            is_tie  = sim_gap < 1e-4

            results.append({
                'phrase':           phrase,
                'n_senses':         n,
                'target':           target,
                'gold_no':          gold_no,
                'gold_def':         clean_text(gold_sense['释义']),
                'pred_no':          pred_no,
                'pred_def':         clean_text(next(
                                        s['释义'] for s in senses if s['释义序号'] == pred_no
                                    )),
                'is_correct':       is_correct,
                'is_tie':           is_tie,
                'sims':             sims,
                'sim_gap':          round(sim_gap, 4),
                'level':            gold_sense.get('等级', ''),
                # ── 消融字段 ──────────────────────────────────────
                'target_matched':   match_info['target_matched'],
                'sense_match_rate': match_info['sense_match_rate'],
            })

            correct += is_correct
            total   += 1
            per_n[n]['total']   += 1
            per_n[n]['correct'] += is_correct

    # ── 消融统计：按 target_matched 分组计算准确率 ────────────────
    ablation: dict = {}
    if any(r['target_matched'] is not None for r in results):
        for status in [True, False]:
            label = 'matched' if status else 'fallback'
            subset = [r for r in results if r['target_matched'] is status]
            if subset:
                n_c = sum(1 for r in subset if r['is_correct'])
                ablation[label] = {
                    'total':    len(subset),
                    'correct':  n_c,
                    'accuracy': round(n_c / len(subset), 4),
                }
        ablation['match_rate'] = round(
            sum(1 for r in results if r['target_matched']) / len(results), 4
        ) if results else 0

    summary = {
        'encoder':  encoder.name,
        'strategy': strategy if not isinstance(encoder, BERTPhraseEncoder)
                    else 'phrase-from-examples (bert_phrase mode)',
        'accuracy': correct / total if total else 0,
        'correct':  correct,
        'total':    total,
        'ties':     sum(1 for r in results if r['is_tie']),
        'per_n':    dict(per_n),
        'ablation': ablation,   # 新增：消融分析结果
    }
    return results, summary


# ══════════════════════════════════════════════════════════════
# 6. 结果写出（不变）
# ══════════════════════════════════════════════════════════════

def _col(ws, idx):
    return get_column_letter(idx)


def write_results(results: list[dict], summary: dict, out_path: str):
    wb = openpyxl.Workbook()

    ws_detail = wb.active
    ws_detail.title = "消歧明细"

    headers = ['短语', '义项数', '学习者例句', '金标序号', '金标释义',
               '预测序号', '预测释义', '是否正确', '是否平局',
               '相似度(各义项)', 'sim差值', '等级',
               '短语定位', '义项例句匹配率']   # 新增消融列
    for col, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEEFF")

    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")

    for row_idx, r in enumerate(results, 2):
        row_data = [
            r['phrase'], r['n_senses'], r['target'],
            r['gold_no'], r['gold_def'],
            r['pred_no'], r['pred_def'],
            '✓' if r['is_correct'] else '✗',
            '△' if r['is_tie'] else '',
            ' | '.join(f"{s:.3f}" for s in r['sims']),
            r['sim_gap'], r['level'],
            # 消融列：None 时显示为空（非 bert_phrase 模式）
            ('✓matched' if r['target_matched'] else '✗fallback')
                if r['target_matched'] is not None else '—',
            f"{r['sense_match_rate']:.0%}"
                if r['sense_match_rate'] is not None else '—',
        ]
        for col, val in enumerate(row_data, 1):
            ws_detail.cell(row=row_idx, column=col, value=val)
        fill = green if r['is_correct'] else red
        for col in range(1, len(headers) + 1):
            ws_detail.cell(row=row_idx, column=col).fill = fill

    for col in range(1, len(headers) + 1):
        ws_detail.column_dimensions[get_column_letter(col)].auto_size = True

    ws_sum = wb.create_sheet("汇总")
    ws_sum.append(['指标', '值'])
    ws_sum.append(['编码器', summary['encoder']])
    ws_sum.append(['义项锚点策略', summary['strategy']])
    ws_sum.append(['总体 Accuracy', f"{summary['accuracy']:.1%}  ({summary['correct']}/{summary['total']})"])
    ws_sum.append(['平局数', summary['ties']])
    ws_sum.append([''])
    ws_sum.append(['义项数', '正确', '总计', 'Accuracy'])
    for n, v in sorted(summary['per_n'].items()):
        acc = v['correct'] / v['total'] if v['total'] else 0
        ws_sum.append([f"{n}义项", v['correct'], v['total'], f"{acc:.1%}"])

    for row in ws_sum.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # ── 消融分析 sheet（仅 bert_phrase 模式有数据）────────────────
    abl = summary.get('ablation', {})
    ws_abl = wb.create_sheet("消融分析")

    if abl:
        # ① 总体匹配率
        ws_abl.append(['=== 短语 token 定位统计 ==='])
        ws_abl.append(['学习者例句定位成功率', f"{abl.get('match_rate', 0):.1%}"])
        ws_abl.append([])

        # ② 按匹配状态分组的准确率（核心消融表）
        ws_abl.append(['=== 消融对比：定位成功 vs 退化为整句向量 ==='])
        ws_abl.append(['分组', '实例数', '正确数', 'Accuracy', '说明'])
        if 'matched' in abl:
            m = abl['matched']
            ws_abl.append([
                '✓ 短语定位成功', m['total'], m['correct'],
                f"{m['accuracy']:.1%}",
                '短语 token 向量有效'
            ])
        if 'fallback' in abl:
            f_ = abl['fallback']
            ws_abl.append([
                '✗ 退化为整句向量', f_['total'], f_['correct'],
                f"{f_['accuracy']:.1%}",
                '词形未匹配，整句 mean-pool 代替'
            ])

        ws_abl.append([])
        ws_abl.append(['解读：若两组 Accuracy 差距明显，说明短语向量方法有效；'
                       '若差距小，说明整句向量已足够，短语定位的额外成本不划算。'])
    else:
        ws_abl.append(['（仅 --encoder bert_phrase 模式下有消融数据）'])

    for row in ws_abl.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
    ws_abl.column_dimensions['A'].width = 28
    ws_abl.column_dimensions['D'].width = 12
    ws_abl.column_dimensions['E'].width = 40

    wb.save(out_path)
    print(f"\n[保存] 结果写入 → {out_path}")


# ══════════════════════════════════════════════════════════════
# 7. 主程序
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='EVP 多义短语消歧 baseline')
    parser.add_argument(
        '--encoder',
        choices=['tfidf', 'bert', 'bert_sentence', 'bert_phrase'],
        default='tfidf',
        help=(
            '编码器类型（默认 tfidf）\n'
            '  tfidf        : TF-IDF，无需联网\n'
            '  bert         : all-MiniLM-L6-v2，句子嵌入小模型（原版）\n'
            '  bert_sentence: bert-base-uncased，整句 mean-pool\n'
            '  bert_phrase  : bert-base-uncased，短语 token 上下文向量'
        )
    )
    parser.add_argument('--strategy', choices=['def', 'def+ex'], default='def+ex',
                        help='义项锚点构建策略（bert_phrase 模式下无效，默认 def+ex）')
    parser.add_argument('--bert-model', default=None,
                        help='自定义模型名（覆盖各编码器默认值，如 bert-large-uncased）')
    parser.add_argument('--spacy-model', default='en_core_web_trf',
                        help='spaCy 模型名，用于 bert_phrase 词形还原')
    parser.add_argument('--data', default='D:\projects\MWED\dab\多义短语.xlsx',
                        help='输入 Excel 路径')
    parser.add_argument('--out',  default='wsd_eval_results.xlsx',
                        help='输出 Excel 路径')
    args = parser.parse_args()

    # ── 若使用默认输出名，自动加上编码器和策略后缀 ──────────────────
    if args.out == 'wsd_eval_results.xlsx':
        if args.encoder == 'bert_phrase':
            # 检查 spaCy 是否可用以决定文件名后缀
            try:
                import spacy as _sp_check
                _sp_check.load(args.spacy_model)
                suffix = 'bert_phrase_lemma'
            except Exception:
                suffix = 'bert_phrase'
        elif args.encoder == 'bert_sentence':
            suffix = f'bert_sentence_{args.strategy}'
        elif args.encoder == 'bert':
            suffix = f'bert_minilm_{args.strategy}'
        else:
            suffix = f'tfidf_{args.strategy}'
        args.out = f'wsd_eval_{suffix}.xlsx'

    # ── 路径统一以脚本所在目录为基准，避免 IDE 工作目录不一致的问题 ──
    _script_dir = Path(__file__).parent
    data_path = (Path(args.data) if Path(args.data).is_absolute()
                 else _script_dir / args.data)
    out_path  = (Path(args.out)  if Path(args.out).is_absolute()
                 else _script_dir / args.out)
    print(f"[INFO] 脚本目录: {_script_dir}")
    print(f"[INFO] 输出文件: {out_path}")

    print(f"[INFO] 读取数据: {data_path}")
    phrase_data  = load_phrase_data(str(data_path))
    eval_phrases = filter_evaluable(phrase_data)
    n_phrases    = len(eval_phrases)
    n_instances  = sum(len(v) for v in eval_phrases.values())
    print(f"[INFO] 可评估: {n_phrases} 条短语，{n_instances} 个义项实例")

    # ── 编码器初始化 ───────────────────────────────────────────────
    if args.encoder == 'bert':
        model = args.bert_model or 'all-MiniLM-L6-v2'
        encoder = BERTEncoder(model)
    elif args.encoder == 'bert_sentence':
        model = args.bert_model or 'bert-base-uncased'
        encoder = BERTSentenceEncoder(model)
    elif args.encoder == 'bert_phrase':
        model = args.bert_model or 'bert-base-uncased'
        encoder = BERTPhraseEncoder(model, spacy_model=args.spacy_model)
    else:
        encoder = TFIDFEncoder()

    print(f"[INFO] 编码器={encoder.name}  策略={args.strategy}")
    print("[INFO] 开始消歧评估...")
    results, summary = evaluate(eval_phrases, encoder, args.strategy)

    print(f"\n{'='*50}")
    print(f"  编码器: {summary['encoder']}")
    print(f"  策略  : {summary['strategy']}")
    print(f"  总体 Accuracy: {summary['accuracy']:.1%}  ({summary['correct']}/{summary['total']})")
    print(f"  平局数: {summary['ties']}")
    print(f"  按义项数:")
    for n, v in sorted(summary['per_n'].items()):
        acc = v['correct'] / v['total'] if v['total'] else 0
        print(f"    {n} 义项: {acc:.1%}  ({v['correct']}/{v['total']})")

    # ── 消融结果打印 ──────────────────────────────────────────────
    abl = summary.get('ablation', {})
    if abl:
        print(f"\n  [消融分析] 短语 token 定位成功率: {abl.get('match_rate', 0):.1%}")
        for group in ['matched', 'fallback']:
            if group in abl:
                g = abl[group]
                label = '✓ 定位成功' if group == 'matched' else '✗ 退化整句'
                print(f"    {label}: {g['accuracy']:.1%}  ({g['correct']}/{g['total']})")
        # 打印 lemma vs exact 命中分布（如有）
        if isinstance(encoder, BERTPhraseEncoder):
            ms = encoder.get_match_stats()
            bm = ms.get('by_method', {})
            if bm:
                print(f"  [匹配方式] ", end='')
                parts = [f"{m}: {v['matched']}条" for m, v in sorted(bm.items())]
                print('  '.join(parts))

    print('='*50)

    write_results(results, summary, str(out_path))


if __name__ == '__main__':
    main()
