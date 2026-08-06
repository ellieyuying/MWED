#!/usr/bin/env python3
"""
消歧金标计分：待标注(人工填的正确义项) + qwen答案，按序号对齐，算消歧准确率。

用法：
    python 消歧金标_计分.py
    默认读同目录 消歧金标_待标注.xlsx（需已填「正确义项」列）+ 消歧金标_qwen答案_勿先看.xlsx

口径：
    单位 = 一条多义命中。
    准确率 = qwen判定义项 == 人工正确义项 的条数 / 已标注条数。
    另单独看 not_applied 这个决策的判准（把"是否 not_applied"当二分类）。
"""
import sys, re
from pathlib import Path
import openpyxl

D = Path(__file__).parent

def load(path, cols):
    wb = openpyxl.load_workbook(path, read_only=True); ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]; I = {x: i for i, x in enumerate(h)}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[I['序号']] is None: continue
        out[int(r[I['序号']])] = {k: (str(r[I[k]]).strip() if r[I[k]] is not None else '') for k in cols if k in I}
    return out

def norm(v):
    v = str(v).strip().lower()
    if v in ('not_applied', 'na', 'n/a', '不适用', '都不合适'): return 'not_applied'
    m = re.search(r'\d+', v)
    return m.group(0) if m else v

def main():
    gold = load(D / '消歧金标_待标注.xlsx', ['序号', '短语', '正确义项'])
    qwen = load(D / '消歧金标_qwen答案_勿先看.xlsx', ['序号', 'qwen判定义项'])

    N = corr = 0; unlabeled = 0
    # not_applied 二分类混淆
    na_tp = na_fp = na_fn = na_tn = 0
    wrong = []
    for i, g in gold.items():
        human = norm(g.get('正确义项', ''))
        if not human or human in ('', 'nan'):
            unlabeled += 1; continue
        q = norm(qwen.get(i, {}).get('qwen判定义项', ''))
        N += 1
        if q == human: corr += 1
        else: wrong.append((i, g.get('短语', ''), f'qwen={q}', f'人工={human}'))
        # not_applied 决策
        qh, hh = (q == 'not_applied'), (human == 'not_applied')
        if qh and hh: na_tp += 1
        elif qh and not hh: na_fp += 1
        elif not qh and hh: na_fn += 1
        else: na_tn += 1

    print(f'已标注 {N} 条（未标 {unlabeled} 条）')
    if N == 0:
        print('还没填「正确义项」列，先标注再跑。'); return
    print(f'\n■ 消歧总准确率 = {corr}/{N} = {corr/N*100:.1f}%')

    print(f'\n■ not_applied 决策（把"是否 not_applied"当二分类）:')
    print(f'   TP{na_tp} FP{na_fp} FN{na_fn} TN{na_tn}')
    if na_tp + na_fp: print(f'   精确率(判na里真na) = {na_tp/(na_tp+na_fp)*100:.1f}%')
    if na_tp + na_fn: print(f'   召回率(真na里判出) = {na_tp/(na_tp+na_fn)*100:.1f}%')

    # 只看"有具体义项"的（排除双方都 not_applied）准确率
    sub = [1 for i, g in gold.items()
           if norm(g.get('正确义项', '')) not in ('', 'nan', 'not_applied')]
    subc = sum(1 for i, g in gold.items()
               if norm(g.get('正确义项', '')) not in ('', 'nan', 'not_applied')
               and norm(qwen.get(i, {}).get('qwen判定义项', '')) == norm(g.get('正确义项', '')))
    if sub: print(f'\n■ 仅"人工判为具体义项"子集的准确率 = {subc}/{len(sub)} = {subc/len(sub)*100:.1f}%')

    print(f'\n■ 判错样例（前 15）:')
    for i, ph, q, h in wrong[:15]:
        print(f'   #{i} 【{ph}】 {q} | {h}')


if __name__ == '__main__':
    main()
