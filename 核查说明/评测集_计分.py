#!/usr/bin/env python3
"""
EVP 短语抽取 —— P/R/F1 计分脚本

用法：
    python 评测集_计分.py [标注后的xlsx路径]
    默认读取同目录的 评测集_短语抽取_待标注.xlsx

口径：
    评测单位 = 一次"短语在某句中的出现"(phrase occurrence)
    TP = 精确率标注表中 判定==1 的条数
    FP = 精确率标注表中 判定==0 的条数
    FN = 召回率标注表中 "漏掉的短语" 列填写的短语总数(按 ; 分隔计数)
    Precision = TP/(TP+FP)   Recall = TP/(TP+FN)   F1 = 2PR/(P+R)

输出：总体指标 + 分层(最少/中位/最多)指标 + 未判定项提醒 + 误命中类型汇总
"""
import sys, re
from pathlib import Path
from collections import defaultdict

import openpyxl


def _cells(ws):
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}
    return idx, list(ws.iter_rows(min_row=2, values_only=True))


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else \
        Path(__file__).parent / '评测集_短语抽取_待标注.xlsx'
    if not path.exists():
        sys.exit(f'找不到文件：{path}')

    wb = openpyxl.load_workbook(path, read_only=True)
    ws_p = wb['精确率标注']
    ws_r = wb['召回率标注']

    # ── 精确率侧 ────────────────────────────────────────────────
    ip, rows_p = _cells(ws_p)
    tp = fp = unjudged = 0
    per_stratum = defaultdict(lambda: {'tp': 0, 'fp': 0, 'fn': 0})
    fp_cases = []
    for r in rows_p:
        v = r[ip['判定(1对/0错)']]
        st = r[ip['分层']]
        if v is None or str(v).strip() == '':
            unjudged += 1
            continue
        s = str(v).strip()
        if s in ('1', '1.0', '对', 'y', 'Y'):
            tp += 1
            per_stratum[st]['tp'] += 1
        elif s in ('0', '0.0', '错', 'n', 'N'):
            fp += 1
            per_stratum[st]['fp'] += 1
            fp_cases.append((r[ip['短语']], r[ip['变体形式']],
                             str(r[ip['句子']])[:60], r[ip['备注']]))
        else:
            unjudged += 1

    # ── 召回率侧 ────────────────────────────────────────────────
    ir, rows_r = _cells(ws_r)
    fn = 0
    fn_cases = []
    for r in rows_r:
        miss = r[ir['漏掉的短语(;分隔)']]
        st = r[ir['分层']]
        if not miss or not str(miss).strip():
            continue
        items = [x.strip() for x in re.split(r'[;；]', str(miss)) if x.strip()]
        fn += len(items)
        per_stratum[st]['fn'] += len(items)
        for it in items:
            fn_cases.append((it, str(r[ir['句子']])[:60]))

    # ── 计分 ────────────────────────────────────────────────────
    def prf(tp, fp, fn):
        p = tp / (tp + fp) if (tp + fp) else 0.0
        rc = tp / (tp + fn) if (tp + fn) else 0.0
        f = 2 * p * rc / (p + rc) if (p + rc) else 0.0
        return p, rc, f

    P, R, F = prf(tp, fp, fn)
    line = '=' * 58
    print(line)
    print('  EVP 短语抽取 —— P/R/F1 评测结果')
    print(line)
    if unjudged:
        print(f'  ⚠ 有 {unjudged} 条尚未判定，未计入（请填完"判定"列）\n')
    print(f'  TP(正确命中) : {tp}')
    print(f'  FP(误命中)   : {fp}')
    print(f'  FN(漏抽)     : {fn}')
    print()
    print(f'  Precision : {P:.1%}   ({tp}/{tp+fp})')
    print(f'  Recall    : {R:.1%}   ({tp}/{tp+fn})')
    print(f'  F1        : {F:.1%}')
    print(line)

    print('\n  分层结果：')
    for st in ('最少', '中位', '最多'):
        d = per_stratum.get(st)
        if not d:
            continue
        p, rc, f = prf(d['tp'], d['fp'], d['fn'])
        print(f'    {st}: P={p:.1%} R={rc:.1%} F1={f:.1%}  '
              f'(TP{d["tp"]} FP{d["fp"]} FN{d["fn"]})')

    if fp_cases:
        print(f'\n  误命中(FP) 明细 —— 共 {len(fp_cases)} 条：')
        agg = defaultdict(int)
        for ph, var, s, note in fp_cases:
            agg[ph] += 1
        for ph, c in sorted(agg.items(), key=lambda x: -x[1])[:15]:
            print(f'    {c:3d}×  {ph}')

    if fn_cases:
        print(f'\n  漏抽(FN) 明细 —— 共 {len(fn_cases)} 条：')
        agg = defaultdict(int)
        for it, s in fn_cases:
            agg[it] += 1
        for it, c in sorted(agg.items(), key=lambda x: -x[1])[:15]:
            print(f'    {c:3d}×  {it}')
    print()


if __name__ == '__main__':
    main()
